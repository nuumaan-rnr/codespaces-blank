"""Compare this app's analysis results with the result tables of an RFEM
export ('COn - 4.1 Members - Internal Forces' sheets) - the validation path
for imported reference models.

Axis mapping: RFEM bends about local y under gravity, this app about local
z, so RFEM My pairs with our Mz and vice versa.  Comparisons use
sign-independent extremes (most compressive N, max |M|) per member so the
different local-axis sign conventions cannot produce false mismatches.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

from .model import RackModel
from .results import CaseResult

KNCM = 1.0e4
KN = 1.0e3

# significance thresholds: values below these (in N / N*mm) are noise
N_THRESHOLD = 0.5 * KN
M_THRESHOLD = 10.0 * KNCM


@dataclass
class MemberRef:
    """RFEM per-member extremes for one combination (app units N, N*mm,
    already mapped to the app's local axes)."""

    N_min: float = 0.0
    N_max: float = 0.0
    My_absmax: float = 0.0
    Mz_absmax: float = 0.0


def read_rfem_results(path: str) -> Dict[str, Dict[int, MemberRef]]:
    """Per combination id ('CO1', ...): member -> extremes."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out: Dict[str, Dict[int, MemberRef]] = {}
    for sheet in wb.sheetnames:
        s = sheet.strip()
        if "4.1 Members" not in s or not s.startswith("CO"):
            continue
        co = s.split(" ")[0]
        members: Dict[int, MemberRef] = {}
        current: Optional[MemberRef] = None
        for r in wb[sheet].iter_rows(min_row=3, values_only=True):
            head = "" if r[0] is None else str(r[0]).strip()
            if head.isdigit():
                current = members.setdefault(int(head), MemberRef())
            if current is None or r[3] is None:
                continue
            try:
                N = float(r[3]) * KN
                My_rfem = float(r[7]) * KNCM
                Mz_rfem = float(r[8]) * KNCM
            except (TypeError, ValueError):
                continue
            current.N_min = min(current.N_min, N)
            current.N_max = max(current.N_max, N)
            # RFEM My -> our Mz, RFEM Mz -> our My
            current.Mz_absmax = max(current.Mz_absmax, abs(My_rfem))
            current.My_absmax = max(current.My_absmax, abs(Mz_rfem))
        out[co] = members
    return out


@dataclass
class Comparison:
    quantity: str
    member: int
    member_set: str
    combo: str
    ours: float
    theirs: float

    @property
    def rel_diff(self) -> float:
        scale = max(abs(self.theirs), abs(self.ours))
        return abs(self.ours - self.theirs) / scale if scale else 0.0


def compare_results(model: RackModel, cases: List[CaseResult],
                    rfem: Dict[str, Dict[int, MemberRef]],
                    skip_members: Tuple[int, ...] = ()) -> List[Comparison]:
    comps: List[Comparison] = []
    for case in cases:
        co = case.combo.split(" ")[0]
        ref = rfem.get(co)
        if ref is None or not case.converged:
            continue
        for mid, mr in case.members.items():
            if mid in skip_members or mid not in ref:
                continue
            m = model.members[mid]
            rm = ref[mid]
            pairs = [("N_min", mr.N_min, rm.N_min, N_THRESHOLD),
                     ("Mz_absmax", mr.Mz_absmax, rm.Mz_absmax, M_THRESHOLD),
                     ("My_absmax", mr.My_absmax, rm.My_absmax, M_THRESHOLD)]
            for q, ours, theirs, thr in pairs:
                if max(abs(ours), abs(theirs)) < thr:
                    continue
                comps.append(Comparison(q, mid, m.member_set, co,
                                        ours, theirs))
    return comps


def summarize(comps: List[Comparison]) -> str:
    if not comps:
        return "No comparable values found."
    lines: List[str] = []
    add = lines.append
    add("# RFEM vs rack15512 (OpenSees) - validation comparison")
    add("")
    diffs = sorted(c.rel_diff for c in comps)

    def pct(p: float) -> float:
        return 100.0 * diffs[min(int(p * len(diffs)), len(diffs) - 1)]
    add(f"- compared values: {len(comps)} "
        f"(member-level extremes of N, My, Mz across combinations)")
    add(f"- median relative difference: {pct(0.50):.1f}%")
    add(f"- 90th percentile: {pct(0.90):.1f}%")
    add(f"- 95th percentile: {pct(0.95):.1f}%")
    add(f"- maximum: {pct(1.0):.1f}%")
    add("")
    add("## By quantity")
    add("")
    add("| quantity | n | median diff | p95 |")
    add("|---|---|---|---|")
    for q in ("N_min", "Mz_absmax", "My_absmax"):
        sub = sorted(c.rel_diff for c in comps if c.quantity == q)
        if not sub:
            continue
        add(f"| {q} | {len(sub)} | {100*sub[len(sub)//2]:.1f}% "
            f"| {100*sub[min(int(0.95*len(sub)), len(sub)-1)]:.1f}% |")
    add("")
    add("## By combination")
    add("")
    add("| combination | n | median diff | p95 |")
    add("|---|---|---|---|")
    combos = sorted({c.combo for c in comps},
                    key=lambda s: (len(s), s))
    for co in combos:
        sub = sorted(c.rel_diff for c in comps if c.combo == co)
        add(f"| {co} | {len(sub)} | {100*sub[len(sub)//2]:.1f}% "
            f"| {100*sub[min(int(0.95*len(sub)), len(sub)-1)]:.1f}% |")
    add("")
    add("## Governing members side by side")
    add("")
    add("For each combination: the most compressed member and the member "
        "with the largest bending moment (by this app's results).")
    add("")
    add("| combination | quantity | member | set | ours | RSTAB/RFEM | diff |")
    add("|---|---|---|---|---|---|---|")
    for co in combos:
        sub = [c for c in comps if c.combo == co]
        for q, div, unit in (("N_min", KN, "kN"), ("Mz_absmax", KNCM, "kNcm")):
            qs = [c for c in sub if c.quantity == q]
            if not qs:
                continue
            gov = max(qs, key=lambda c: abs(c.ours))
            add(f"| {co} | {q} | {gov.member} | {gov.member_set[:18]} "
                f"| {gov.ours/div:.2f} {unit} | {gov.theirs/div:.2f} {unit} "
                f"| {100*gov.rel_diff:.1f}% |")
    add("")
    add("### Reading the tail")
    add("")
    add("Large *relative* differences concentrate in members with tiny "
        "*absolute* forces (sub-kN bracing axials, moments of a few kNcm) "
        "where the imperfection conventions differ: this app applies "
        "equivalent horizontal forces at every loaded node, RFEM applies "
        "inclinations to the upright member sets.  Check the absolute "
        "values in the table below before treating a row as a "
        "discrepancy.")
    add("")
    add("## Largest differences")
    add("")
    add("| member | set | combo | quantity | ours | RFEM | diff |")
    add("|---|---|---|---|---|---|---|")
    worst = sorted(comps, key=lambda c: -c.rel_diff)[:20]
    for c in worst:
        div, unit = (KN, "kN") if c.quantity.startswith("N") else (KNCM, "kNcm")
        add(f"| {c.member} | {c.member_set[:18]} | {c.combo} | {c.quantity} "
            f"| {c.ours/div:.2f} {unit} | {c.theirs/div:.2f} {unit} "
            f"| {100*c.rel_diff:.1f}% |")
    return "\n".join(lines)
