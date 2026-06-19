"""Generate the consolidated section-master Excel template.

One workbook, four sheets, that captures EVERYTHING needed to import a company
master in a single file:

  * README          - units, axis convention and how each sheet is used
  * SECTIONS        - one row per section (uprights, beams, bracings, others)
                      with the full property set
  * BASE_STIFFNESS  - per-upright floor-connection table (N vs k_b vs M_Rd)
  * BEAM_STIFFNESS  - per-beam connector stiffness Kb by upright wall thickness

Fill it in and import it once (rack15512.master_xlsx.load_master detects the
SECTIONS sheet); the importer converts the labelled units to the model's N/mm.
"""

from __future__ import annotations

# SECTIONS columns: (header, example upright, example beam, example bracing,
# example other) - the header carries the unit the importer expects.
_SECTION_COLUMNS = [
    ("name", "UP0002", "RHS 60x40x1.6", "1C36x21x1.5", "DRIVEIN2.5"),
    ("role", "upright", "beam", "bracing", "others"),
    ("description", "Upright 90x61x1.6", "RHS beam", "lipped channel brace",
     "drive-in rail"),
    ("fy (MPa)", 355, 275, 355, 235),
    ("A (cm2)", 3.18, 3.03, 0.99, 5.59),
    ("Iy (cm4) [minor / cross-aisle]", 1.06, 8.15, 0.71, 9.36),
    ("Iz (cm4) [major / down-aisle]", 3.37, 15.21, 1.05, 20.75),
    ("J (cm4)", 0.02, 16.94, 0.02, 0.8),
    ("Wely (cm3) [minor]", 0.27, 4.07, 0.41, 2.94),
    ("Welz (cm3) [major]", 0.75, 5.07, 0.54, 4.0),
    ("Avy (cm2) [local-y shear]", 0.44, 0.92, 0.39, 1.17),
    ("Avz (cm2) [local-z shear]", 0.82, 1.68, 0.17, 2.58),
    ("It_gross (cm4)", 0.02, 16.94, 0.02, 0.8),
    ("Iw_gross (cm6)", 308.31, 1.41, 0.0, 0.0),
    ("y0 (cm) [shear-centre offset]", 4.59, 0.0, 0.0, 0.0),
    ("mount_offset (mm) [stiffener centroid gap to upright]", "", "", "", 30),
    ("depth_h (mm)", 90, 60, 36, 159),
    ("width_b (mm)", 61, 40, 21, 128),
    ("t (mm)", 1.6, 1.6, 1.5, 2.5),
    ("e1 (mm)", 13.48, "", "", ""),
    ("e2 (mm)", 12.39, "", "", ""),
    ("curve_y", "c", "c", "c", "c"),
    ("curve_z", "c", "c", "c", "c"),
    ("connector_k (kNm/rad)", "", 200, "", ""),
    ("connector_m_rd (kNm)", "", 1.14, "", ""),
]

_README = [
    "Section-master template — fill one file and import it once.",
    "",
    "Sheets:",
    "  SECTIONS        one row per section; role = upright / beam / bracing / others",
    "  BASE_STIFFNESS  per-upright floor connection: N (kN), k_b (kNcm/rad), M_Rd (kNcm)",
    "  BEAM_STIFFNESS  per-beam connector: M_Rd (kNcm) and Kb @ UPL <t> (kNcm/rad)",
    "",
    "Units are in each SECTIONS column header (cm2/cm4/cm3/cm6, mm, MPa, kNm/rad);",
    "the importer converts to the solver's N/mm internally.",
    "",
    "Axis convention (IMPORTANT): local z = MAJOR / down-aisle (the strong axis",
    "engaged by down-aisle bending of uprights and gravity bending of beams);",
    "local y = MINOR / cross-aisle. Put the larger second moment in 'Iz'.",
    "",
    "Optional columns may be left blank (they default to gross / safe values).",
    "Avy/Avz enable shear-flexible (Timoshenko) members; It_gross/Iw_gross/y0",
    "enable the flexural-torsional buckling check (uprights).",
    "mount_offset (stiffener rows only): cross-aisle distance from the upright",
    "centroid line to the stiffener centroid line when mounted; the builder uses",
    "it to place the stiffener node (else the app's global stiffener offset).",
    "",
    "BEAM_STIFFNESS: add a 'Kb @ UPL x.x' column per upright wall thickness; the",
    "beam connector stiffness is then chosen by the upright the beam bolts to.",
    "A company name is requested on import (sections are company-specific).",
]

_BASE_COLUMNS = ["upright", "N (kN)", "k_b (kNcm/rad)", "M_Rd (kNcm)"]
_BASE_EXAMPLE = [["UP0002", 30, 1823, 142], ["UP0002", 40, 3470, 173],
                 ["UP0002", 50, 5117, 193], ["UP0002", 60, 6764, 203]]

_BEAM_COLUMNS = ["Section", "M_Rd (kNcm)", "Kb @ UPL 1.6 (kNcm/rad)",
                 "Kb @ UPL 2.0 (kNcm/rad)", "Kb @ UPL 2.5 (kNcm/rad)"]
_BEAM_EXAMPLE = [["RHS 60x40x1.6", 114, 1566, 2038.5, 2492],
                 ["RHS 80x40x1.6", 152, 2382.4, 2726.4, 2942]]


def build_master_template(path: str) -> str:
    """Write the consolidated section-master template workbook to `path`."""
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    bold = Font(bold=True)

    readme = wb.active
    readme.title = "README"
    for i, line in enumerate(_README, start=1):
        readme.cell(i, 1, line)
    readme.column_dimensions["A"].width = 90

    sec = wb.create_sheet("SECTIONS")
    for c, col in enumerate(_SECTION_COLUMNS, start=1):
        sec.cell(1, c, col[0]).font = bold
        for j in range(4):                       # 4 example rows
            sec.cell(2 + j, c, col[1 + j])
        sec.column_dimensions[sec.cell(1, c).column_letter].width = \
            max(12, len(col[0]) + 2)

    base = wb.create_sheet("BASE_STIFFNESS")
    for c, h in enumerate(_BASE_COLUMNS, start=1):
        base.cell(1, c, h).font = bold
    for r, row in enumerate(_BASE_EXAMPLE, start=2):
        for c, v in enumerate(row, start=1):
            base.cell(r, c, v)

    beam = wb.create_sheet("BEAM_STIFFNESS")
    for c, h in enumerate(_BEAM_COLUMNS, start=1):
        beam.cell(1, c, h).font = bold
    for r, row in enumerate(_BEAM_EXAMPLE, start=2):
        for c, v in enumerate(row, start=1):
            beam.cell(r, c, v)

    wb.save(path)
    return path


def template_bytes() -> bytes:
    """Return the template workbook as bytes (for a download button)."""
    import io
    import openpyxl  # noqa: F401  (ensure the dependency is present)
    import tempfile
    import os
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    build_master_template(tmp.name)
    with open(tmp.name, "rb") as f:
        data = f.read()
    os.unlink(tmp.name)
    return data
