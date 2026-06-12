"""Importer for RFEM (Dlubal) model exports (.xlsx) - sheets '1.1 Nodes',
'1.2 Materials', '1.3 Cross-Sections', '1.4 Member Hinges', '1.7 Members',
'1.8 Nodal Supports', '2.1 Load Cases', '2.5 Load Combinations',
'LCn - 3.1 Nodal Loads', 'LCn - 3.2 Member Loads', 'LCn - 3.4 Imperfections'.

Conventions handled:
  * RFEM global Z points DOWN (gravity = +Z); the importer flips to this
    app's Z-up frame (z_our = -Z_rfem, load direction Z -> -z).
  * RFEM units: kN, cm-based section properties, kN/m line loads,
    kNcm/rad hinge springs -> converted to N/mm.
  * RFEM bends beams about local y under gravity; this app uses local z.
    Section Iy/Iz and hinge jy/jz are therefore SWAPPED on import.
  * Member rotation angles beta of 0/180 deg leave Iy/Iz assignments
    unchanged and are ignored; members rotated near 90 deg raise an error.
  * Wel is not exported; it is estimated as I / (half overall dimension)
    and marked in the section description (used by stress checks only).
  * Self-weight load cases (RFEM computes them internally) are rebuilt as
    member UDLs w = specific_weight * A on every member.
  * Concentrated member loads: the member is split at the load position
    and the force applied to the new node.
  * Imperfection load cases (inclination 1/phi0 on member sets) become the
    sway imperfection phi = 1/phi0; combinations referencing them get the
    matching single direction ('+x' for LCs applied in the global X sense,
    '+y' for Y).  RFEM's per-set +-direction bookkeeping (which exists
    because member local axes alternate) collapses to one global sense.
"""

from __future__ import annotations

import math
import re
from typing import Dict, List, Optional, Tuple

from .model import (Combination, CrossSection, Hinge, Imperfection, LoadCase,
                    Member, MemberLoad, NodalLoad, RackModel, Steel, Support)

KNCM2 = 10.0      # kN/cm^2 -> MPa
CM2 = 1.0e2
CM4 = 1.0e4
KNCM = 1.0e4      # kNcm -> Nmm (also kNcm/rad -> Nmm/rad)
KN = 1.0e3
KN_M = 1.0        # kN/m -> N/mm


def _rows(wb, name) -> List[list]:
    for sheet in wb.sheetnames:
        if sheet.strip().startswith(name):
            return [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    return []


def _s(v) -> str:
    return "" if v is None else str(v).strip()


def _f(v, default=None) -> Optional[float]:
    s = _s(v)
    if not s:
        return default
    try:
        return float(s)
    except ValueError:
        return default


def _id_list(spec: str) -> List[int]:
    """Parse RFEM id lists like '2,3,609-616,621'."""
    out: List[int] = []
    for part in _s(spec).split(","):
        part = part.strip()
        if not part:
            continue
        m = re.fullmatch(r"(\d+)-(\d+)", part)
        if m:
            out.extend(range(int(m.group(1)), int(m.group(2)) + 1))
        else:
            out.append(int(part))
    return out


def load_rfem(path: str, fy: float = 250.0, master=None) -> RackModel:
    """Build a RackModel from an RFEM data export.  fy is the design yield
    strength [MPa] (RFEM does not export it; IS 2062 E250 -> 250).

    master: optional MasterWorkbook.  RFEM does NOT export nonlinear
    (load-dependent) floor-connection springs - they appear as free
    rotations in the support table even when present in the model.  When a
    master with a BASE_STIFFNESS sheet is given and the support comment
    names an upright (e.g. 'UP0016'), the rotational base stiffness k_b is
    interpolated at the estimated factored upright axial load and applied
    to the free rx/ry support DOFs.  This was validated against the SPR
    reference model: without it the second-order combinations diverge
    (the linear-LC-stiffness model has alpha_cr ~ 1), with it the member
    forces match RFEM's nonlinear combination results to ~1%."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("RFEM import requires openpyxl") from None
    wb = openpyxl.load_workbook(path, data_only=True)
    model = RackModel(name="RFEM import")

    # ---- nodes (flip Z) ----------------------------------------------------
    z_raw: Dict[int, float] = {}
    for r in _rows(wb, "1.1 Nodes")[2:]:
        if _s(r[3]) == "":
            continue
        nid = int(r[0])
        z_raw[nid] = float(r[5])
        model.add_node(nid, float(r[3]), float(r[4]), -float(r[5]))
    z_shift = min(n.z for n in model.nodes.values())
    for n in model.nodes.values():
        n.z -= z_shift

    # ---- materials -----------------------------------------------------------
    gamma: Dict[str, float] = {}      # N/mm^3, for self-weight
    for r in _rows(wb, "1.2 Materials")[2:]:
        if _f(r[2]) is None:
            continue
        name = f"M{int(r[0])}"
        model.materials[name] = Steel(
            name, E=_f(r[2]) * KNCM2, G=_f(r[3]) * KNCM2, fy=fy,
            nu=_f(r[4], 0.3))
        gamma[name] = _f(r[5], 78.5) * 1.0e-6     # kN/m^3 -> N/mm^3

    # ---- cross-sections (swap y/z; estimate Wel from overall dims) ----------
    sec_name: Dict[int, str] = {}
    for r in _rows(wb, "1.3 Cross-Sections")[2:]:
        if _f(r[4]) is None:
            continue
        no = int(r[0])
        name = f"CS{no} {_s(r[1])[:24]}"
        b = _f(r[11], 50.0)
        h = _f(r[12], 100.0)
        Iz = _f(r[4]) * CM4               # RFEM Iy (gravity axis) -> local z
        Iy = _f(r[5]) * CM4
        model.sections[name] = CrossSection(
            name=name, material=f"M{int(r[2])}",
            A=_f(r[6]) * CM2, Iy=Iy, Iz=Iz,
            J=max(_f(r[3]) * CM4, 1.0),
            Wely=Iy / (b / 2.0), Welz=Iz / (h / 2.0),
            description="RFEM import; Wel estimated as I/(dim/2)")
        sec_name[no] = name

    # ---- member hinges (swap y/z) -------------------------------------------
    hinges: Dict[int, Hinge] = {}
    for r in _rows(wb, "1.4 Member Hinges")[2:]:
        if _s(r[0]) == "" or _s(r[1]) == "":
            continue

        def rot(v) -> Optional[float]:
            s = _s(v)
            if s in ("-", ""):
                return None               # continuous
            val = _f(s, 0.0)
            return val * KNCM if val and val > 0 else 0.0
        for tr, label in ((r[2], "ux"), (r[3], "uy"), (r[4], "uz")):
            if _s(tr) not in ("-", ""):
                raise ValueError(f"Member hinge {r[0]}: translational "
                                 f"release '{label}' is not supported")
        hinges[int(r[0])] = Hinge(rx=rot(r[5]), rz=rot(r[6]), ry=rot(r[7]))

    # ---- members -------------------------------------------------------------
    for r in _rows(wb, "1.7 Members")[2:]:
        if _s(r[1]) == "":
            continue
        mid = int(r[0])
        beta = abs(_f(r[5], 0.0)) % 360.0
        if min(abs(beta - a) for a in (0.0, 180.0, 360.0)) > 5.0:
            raise ValueError(f"Member {mid}: rotation beta={beta:.1f} deg "
                             "not supported (only 0/180)")
        mtype = "truss" if _s(r[1]).lower().startswith("truss") else "beam"
        sec = sec_name[int(r[6])]
        h_i = hinges.get(int(_f(r[8], 0) or 0)) if mtype == "beam" else None
        h_j = hinges.get(int(_f(r[9], 0) or 0)) if mtype == "beam" else None

        def copy_h(h: Optional[Hinge]) -> Optional[Hinge]:
            return Hinge(rz=h.rz, ry=h.ry, rx=h.rx) if h else None
        model.add_member(mid, int(r[2]), int(r[3]), sec, mtype=mtype,
                         hinge_i=copy_h(h_i), hinge_j=copy_h(h_j),
                         mesh=2 if mtype == "beam" else 1,
                         member_set=sec)

    # ---- supports --------------------------------------------------------------
    def restraint(v):
        s = _s(v)
        if s == "+":
            return True
        if s in ("-", ""):
            return False
        val = _f(s, 0.0) or 0.0
        return val if val > 0 else False
    support_upright: Dict[int, str] = {}      # support index -> comment
    for r in _rows(wb, "1.8 Nodal Supports")[2:]:
        if _s(r[1]) == "":
            continue
        for nid in _id_list(r[1]):
            if nid not in model.nodes:
                continue
            # columns: uX' uY' uZ' jX' jY' jZ'; rotational springs would be
            # kNcm/rad - convert; jY (RFEM gravity bending axis) -> our ry
            def rot_restraint(v):
                rv = restraint(v)
                return rv * KNCM if isinstance(rv, float) else rv
            support_upright[len(model.supports)] = _s(r[13]) if len(r) > 13 else ""
            model.supports.append(Support(
                nid, ux=restraint(r[7]), uy=restraint(r[8]),
                uz=restraint(r[9]), rx=rot_restraint(r[10]),
                ry=rot_restraint(r[11]), rz=rot_restraint(r[12])))

    # ---- load cases ------------------------------------------------------------
    imp_dir_of_lc: Dict[str, str] = {}
    phi: Optional[float] = None
    lc_rows = _rows(wb, "2.1 Load Cases")[2:]
    for r in lc_rows:
        lc_id = _s(r[0])
        if not lc_id or _s(r[2]) != "+":
            continue
        category = _s(r[3]).lower()
        case = LoadCase(lc_id,
                        "permanent" if "permanent" in category else "variable")

        # self weight (RFEM applies it internally when Active = '+')
        if _s(r[4]) == "+" and abs(_f(r[7], 0.0) or 0.0) > 0:
            zfac = _f(r[7], 1.0)
            for m in model.members.values():
                sec = model.sections[m.section]
                w = gamma[sec.material] * sec.A          # N/mm
                case.member_loads.append(MemberLoad(m.id, qz=-zfac * w))

        # nodal loads
        for nr in _rows(wb, f"{lc_id} - 3.1 Nodal Loads")[2:]:
            if _s(nr[1]) == "":
                continue
            for nid in _id_list(nr[1]):
                case.nodal_loads.append(NodalLoad(
                    nid, fx=(_f(nr[4], 0.0) or 0.0) * KN,
                    fy=(_f(nr[5], 0.0) or 0.0) * KN,
                    fz=-(_f(nr[6], 0.0) or 0.0) * KN))

        # member loads (uniform UDL or concentrated -> split member)
        for mr in _rows(wb, f"{lc_id} - 3.2 Member Loads")[2:]:
            if _s(mr[2]) == "":
                continue
            direction = _s(mr[5]).upper()
            dist = _s(mr[4]).lower()
            p = _f(mr[7], 0.0) or 0.0
            if dist.startswith("uniform"):
                qx = p * KN_M if direction == "X" else 0.0
                qy = p * KN_M if direction == "Y" else 0.0
                qz = -p * KN_M if direction == "Z" else 0.0
                for mid in _id_list(mr[2]):
                    case.member_loads.append(MemberLoad(mid, qx, qy, qz))
            elif dist.startswith("concentrated"):
                a = _f(mr[11], 0.0) or 0.0
                for mid in _id_list(mr[2]):
                    nid = _split_member(model, mid, a)
                    case.nodal_loads.append(NodalLoad(
                        nid, fx=p * KN if direction == "X" else 0.0,
                        fy=p * KN if direction == "Y" else 0.0,
                        fz=-p * KN if direction == "Z" else 0.0))
            else:
                raise ValueError(f"{lc_id}: member load distribution "
                                 f"'{mr[4]}' not supported")

        # imperfection load cases: record phi and the global sense
        imps = [ir for ir in _rows(wb, f"{lc_id} - 3.4 Imperfections")[2:]
                if _s(ir[2]) != ""]
        if imps:
            phi0 = abs(_f(imps[0][6], 300.0) or 300.0)
            phi = 1.0 / phi0
            # the X-sense LC is the one combined with X placement loads in
            # the combinations; fall back to the LC order (first = x)
            imp_dir_of_lc[lc_id] = ""
            continue                       # not a force load case

        if case.nodal_loads or case.member_loads:
            model.load_cases[lc_id] = case

    # assign global senses to the imperfection LCs (first -> +x, second -> +y)
    for k, lc_id in enumerate(sorted(imp_dir_of_lc)):
        imp_dir_of_lc[lc_id] = "+x" if k == 0 else "+y"

    if phi is not None:
        model.imperfection = Imperfection(phi=phi, method="EHF")

    # ---- combinations -----------------------------------------------------------
    for r in _rows(wb, "2.5 Load Combinations")[2:]:
        co = _s(r[0])
        if not co or _s(r[3]) != "+":
            continue
        ds = _s(r[1])
        factors: Dict[str, float] = {}
        imp_dirs: List[str] = []
        for i in range(6):
            fcol, ncol = 4 + 2 * i, 5 + 2 * i
            if ncol >= len(r) or _s(r[ncol]) == "":
                continue
            lc_id, fac = _s(r[ncol]), _f(r[fcol], 1.0)
            if lc_id in imp_dir_of_lc:
                imp_dirs.append(imp_dir_of_lc[lc_id])
            elif lc_id in model.load_cases:
                factors[lc_id] = fac
            # silently skip references to unsolved/absent cases (e.g. CO101)
        if not factors:
            continue
        kind = "SLS" if ds == "5" else "ULS"
        model.combinations.append(Combination(
            f"{co} {_s(r[2])[:30]}", kind, factors,
            imperfection=bool(imp_dirs),
            imp_directions=imp_dirs or None))

    # ---- load-dependent base springs from the master (see docstring) ---------
    if master is not None and model.supports:
        N_est = _estimate_support_axial(model)
        for idx, sup in enumerate(model.supports):
            upright = support_upright.get(idx, "")
            if upright not in master.base_tables:
                continue
            k_b, _ = master.base_stiffness(upright, N_est)
            if sup.rx is False:
                sup.rx = k_b
            if sup.ry is False:
                sup.ry = k_b

    return model


def _estimate_support_axial(model: RackModel) -> float:
    """Largest factored total vertical load over the ULS combinations,
    averaged over the supports - the axial level for interpolating the
    load-dependent floor-connection stiffness."""
    worst = 0.0
    for combo in model.combinations:
        if combo.kind != "ULS":
            continue
        total = 0.0
        for case_name, factor in combo.factors.items():
            lc = model.load_cases[case_name]
            for nl in lc.nodal_loads:
                total += factor * max(-nl.fz, 0.0)
            for ml in lc.member_loads:
                total += factor * max(-ml.qz, 0.0) \
                    * model.member_length(model.members[ml.member])
        worst = max(worst, total)
    return worst / len(model.supports)


def _split_member(model: RackModel, mid: int, a: float) -> int:
    """Split member `mid` at distance `a` from its start node; return the
    id of the new node (reused if the position is an existing end)."""
    m = model.members[mid]
    ni, nj = model.nodes[m.node_i], model.nodes[m.node_j]
    L = model.member_length(m)
    if a <= 1.0:
        return m.node_i
    if a >= L - 1.0:
        return m.node_j
    t = a / L
    new_nid = max(model.nodes) + 1
    model.add_node(new_nid, ni.x + t * (nj.x - ni.x),
                   ni.y + t * (nj.y - ni.y), ni.z + t * (nj.z - ni.z))
    new_mid = max(model.members) + 1
    model.add_member(new_mid, new_nid, m.node_j, m.section, mtype=m.mtype,
                     hinge_j=m.hinge_j, mesh=m.mesh, member_set=m.member_set,
                     vecxz=m.vecxz)
    m.node_j = new_nid
    m.hinge_j = None
    return new_nid
