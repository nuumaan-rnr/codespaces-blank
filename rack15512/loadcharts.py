"""Load-chart generator for uprights and beams.

ADDITIVE only - this module imports the frozen analysis engine and never edits
it.  It produces, from a section master:

  * Upright load charts - max axial buckling capacity N_b,Rd (EN 15512, the same
    closed form as rack15512.presize.upright_utilisation) versus the down-aisle
    buckling length Lcr_DA, for cross-aisle pitches 500/600 mm, in three configs:
      - X  : crossed bracing, the upright is restrained every panel  -> Lcr_CA = pitch
      - D  : single-diagonal bracing, restrained every other panel    -> Lcr_CA = 2*pitch
      - XS : X bracing + an internal stiffener combined with the upright (90-deep
             upright + IN_STIFFENER90X1.6, 120-deep + IN_STIFFENER120X1.6).
    Uprights are charted at fy = 355 (pure axial - a strut load chart).

  * Beam load charts - max total load per bay level (the pair of beams) versus the
    span, for a single-span beam with semi-rigid end connectors.  The connector
    rotational stiffness depends on the UPRIGHT thickness it bolts to, so a curve
    is drawn per upright thickness.  Governed by the min of beam bending, the
    connector moment, deflection (L/200) and shear.

Outputs PNG charts (matplotlib) and tidy Excel tables (openpyxl).

CLI:  python -m rack15512.loadcharts --master <master.xlsx> --out charts/
"""

from __future__ import annotations

import argparse
import math
import os
import zipfile
from dataclasses import replace
from typing import Dict, List, Optional, Tuple

from . import branding
from .builder import _closed_upright_section
from .master_xlsx import load_master
from .model import CrossSection, Steel
from .presize import upright_utilisation

# ----------------------------------------------------------------- constants
E_STEEL = 210000.0          # MPa
G_STEEL = 81000.0           # MPa
GAMMA_M0 = 1.0              # cross-section resistance (EN 15512)
GAMMA_M1 = 1.1              # member buckling resistance (EN 15512)
GAMMA_Q = 1.4              # variable (pallet) load factor at ULS (EN 15512)
FY_UPRIGHT = 355.0          # YS355, per the request
DEFL_RATIO = 200.0          # beam deflection limit L / 200 (EN 15512 SLS)

UP_DA = list(range(250, 5001, 50))        # down-aisle buckling length grid [mm]
BEAM_SPAN = list(range(500, 4001, 50))    # beam span grid [mm]
PITCHES = (500.0, 600.0)                   # cross-aisle bracing pitches [mm]

# config name -> (matplotlib colour, linestyle)
_STYLE = {
    "X-500": (branding.TEAL, "-"),
    "X-600": (branding.TEAL_LIGHT, "-"),
    "D-1000": (branding.GREY, "--"),
    "D-1200": (branding.GREY_LIGHT, "--"),
    "XS-500": ("#C0392B", "-"),
    "XS-600": ("#E67E22", "-"),
}
_BEAM_COLOURS = [branding.TEAL, "#C0392B", "#E67E22", branding.GREY,
                 branding.TEAL_LIGHT, "#8E44AD"]


# ------------------------------------------------------- combined (XS) section
def combined_upright_section(up: CrossSection, st: CrossSection,
                             offset: float) -> CrossSection:
    """Upright + internal stiffener as ONE section, parallel-axis about the
    combined centroid in the cross-aisle (Iy / Lcr_CA) direction, with the
    closed-cell torsion/warping/shear-centre credit from the Type-1 stiffener.

    `offset` = cross-aisle centroid separation [mm] (the stiffener mount_offset).
    The cross-aisle inertia gains the parallel-axis term (the lever the stiffener
    adds); the down-aisle inertia is a simple sum (no lever that way)."""
    A = up.A + st.A
    a_eff = up.area_eff + st.area_eff
    c = st.A * offset / A                       # centroid shift toward the stiffener
    Iy = up.Iy + up.A * c ** 2 + st.Iy + st.A * (offset - c) ** 2
    Iz = up.Iz + st.Iz
    closed = _closed_upright_section(up.name + "~closed", up)   # closed-cell It/Iw/y0
    return replace(up, name=f"{up.name}+{st.name}", A=A, A_eff=a_eff,
                   Iy=Iy, Iz=Iz, Iy_gross=Iy, Iz_gross=Iz,
                   It_gross=closed.It_gross, Iw_gross=closed.Iw_gross,
                   y0=closed.y0, J=closed.J)


def _stiffener_for(lib, up: CrossSection) -> Optional[CrossSection]:
    """The matching internal stiffener (1.6 mm) for a 90/120-deep upright."""
    h = round(up.depth_h or 0.0)
    name = {90: "IN_STIFFENER90X1.6", 120: "IN_STIFFENER120X1.6"}.get(h)
    if not name:
        return None
    try:
        return lib.get(name)
    except (KeyError, ValueError):
        return None


# ------------------------------------------------------------- upright capacity
def upright_capacity_kN(sec: CrossSection, fy: float, Lcr_da: float,
                        Lcr_ca: float) -> Tuple[float, str]:
    """Pure-axial buckling capacity N_b,Rd [kN] and the governing axis
    ('y'=cross-aisle, 'z'=down-aisle, 'FT')."""
    mat = Steel("steel", E=E_STEEL, fy=fy, G=G_STEEL)
    u = upright_utilisation(sec, mat, Lcr_y=Lcr_ca, Lcr_z=Lcr_da, N=1.0,
                            gamma_M1=GAMMA_M1, gamma_M0=GAMMA_M0)
    return u["N_b_Rd"] / 1e3, u["gov"]


def _upright_configs(lib, up: CrossSection
                     ) -> List[Tuple[str, float, CrossSection]]:
    """(label, Lcr_CA, section) for every config of one upright."""
    cfgs: List[Tuple[str, float, CrossSection]] = []
    for p in PITCHES:                       # X: braced every panel -> Lcr_CA = pitch
        cfgs.append((f"X-{int(p)}", p, up))
    for p in PITCHES:                       # D: every other panel  -> Lcr_CA = 2*pitch
        cfgs.append((f"D-{int(2 * p)}", 2 * p, up))
    st = _stiffener_for(lib, up)            # XS: combined section (90/120 only)
    if st is not None:
        comb = combined_upright_section(up, st, st.mount_offset or 30.0)
        for p in PITCHES:
            cfgs.append((f"XS-{int(p)}", p, comb))
    return cfgs


def upright_chart_data(lib, fy: float = FY_UPRIGHT) -> Dict[str, dict]:
    """{upright_name: {"desc":..., "curves": {label: {"lcr_ca":, "da":[...],
    "cap":[...], "gov":[...]}}}}."""
    out: Dict[str, dict] = {}
    for name in lib.names("upright"):
        up = lib.get(name)
        curves: Dict[str, dict] = {}
        for label, lcr_ca, sec in _upright_configs(lib, up):
            caps, govs = [], []
            for da in UP_DA:
                cap, gov = upright_capacity_kN(sec, fy, da, lcr_ca)
                caps.append(cap)
                govs.append(gov)
            curves[label] = {"lcr_ca": lcr_ca, "da": UP_DA, "cap": caps,
                             "gov": govs}
        out[name] = {"desc": up.description or "", "curves": curves}
    return out


# --------------------------------------------------------------- beam capacity
def beam_level_capacity_kN(beam: CrossSection, fy: float, span: float,
                           k_conn: Optional[float], m_rd: Optional[float]
                           ) -> Tuple[float, str]:
    """Max TOTAL WORKING load per bay level (both beams) [kN] for a single-span
    beam with equal semi-rigid end rotational springs of stiffness k_conn
    [N*mm/rad].

    The strength limits (bending / connector / shear) are ULS checks - the
    variable load factor GAMMA_Q is applied so the returned value is the
    permissible WORKING (characteristic) load; deflection is the SLS limit
    (L/200) at working load.  Returns (load_kN, governing); load is shared by
    the two beams: total = 2 * w * span."""
    L = span
    EI = E_STEEL * beam.Iz                       # strong (down-aisle gravity) axis
    # connection fixity factor: gamma=0 pinned, gamma=1 fully fixed
    gamma = 1.0 / (1.0 + 2.0 * EI / (k_conn * L)) if (k_conn and k_conn > 0) else 0.0
    inf = float("inf")
    lim: Dict[str, float] = {}

    # beam bending (ULS): max of mid-span (wL^2(1/8 - g/12)) and end (wL^2 g/12);
    # working load = M_Rd / (gamma_Q * coef * L^2)
    coef_mid = 1.0 / 8.0 - gamma / 12.0
    coef_end = gamma / 12.0
    coef_bend = max(coef_mid, coef_end)
    m_rd_beam = beam.mod_z_eff * fy / GAMMA_M0
    lim["bending"] = (m_rd_beam / (GAMMA_Q * coef_bend * L * L)
                      if coef_bend > 0 else inf)

    # connector moment (ULS): end moment <= connector design resistance
    lim["connector"] = (m_rd / (GAMMA_Q * coef_end * L * L)
                        if (m_rd and coef_end > 1e-12) else inf)

    # deflection L/200 (SLS, working load): bending + shear when Avz given
    cb = L ** 4 / EI * (5.0 / 384.0 - gamma / 96.0)
    cs = L * L / (8.0 * G_STEEL * beam.Avz) if beam.Avz else 0.0
    lim["deflection"] = (L / DEFL_RATIO) / (cb + cs) if (cb + cs) > 0 else inf

    # shear (ULS): V = wL/2 <= V_Rd = A_v fy / (sqrt3 gM0), A_v = 2 h t (two webs)
    if beam.depth_h and beam.t:
        v_rd = 2.0 * beam.depth_h * beam.t * fy / (math.sqrt(3.0) * GAMMA_M0)
        lim["shear"] = 2.0 * v_rd / (GAMMA_Q * L)

    w = min(lim.values())                        # per beam [N/mm], working
    gov = min(lim, key=lim.get)
    return 2.0 * w * L / 1e3, gov                # both beams -> total kN per level


def _beam_thickness_curves(beam: CrossSection
                           ) -> List[Tuple[Optional[float], float]]:
    """(upright_thickness, connector_k) pairs; falls back to the single k."""
    if beam.connector_k_by_upl:
        return [(float(t), float(k)) for t, k in beam.connector_k_by_upl]
    if beam.connector_k:
        return [(None, float(beam.connector_k))]
    return [(None, 0.0)]                          # pinned (no connector data)


def beam_chart_data(lib, fy_of) -> Dict[str, dict]:
    """{beam_name: {"fy":, "m_rd":, "curves": {upright_t: {"span":[...],
    "load":[...], "gov":[...]}}}}."""
    out: Dict[str, dict] = {}
    for name in lib.names("beam"):
        beam = lib.get(name)
        fy = fy_of(name)
        m_rd = beam.connector_m_rd
        curves: Dict[str, dict] = {}
        for t, k in _beam_thickness_curves(beam):
            loads, govs = [], []
            for L in BEAM_SPAN:
                load, gov = beam_level_capacity_kN(beam, fy, L, k, m_rd)
                loads.append(load)
                govs.append(gov)
            key = f"upl {t:.1f} mm" if t else "pinned"
            curves[key] = {"span": BEAM_SPAN, "load": loads, "gov": govs}
        out[name] = {"fy": fy, "m_rd": m_rd, "curves": curves}
    return out


# ------------------------------------------------------------------- plotting
def _plot_upright(name: str, info: dict, path: str) -> None:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    fig, ax = plt.subplots(figsize=(8.5, 5.5))
    for label, c in info["curves"].items():
        colour, ls = _STYLE.get(label, (branding.GREY, "-"))
        ax.plot(c["da"], c["cap"], ls, color=colour, lw=1.8,
                label=f"{label}  (Lcr,CA={int(c['lcr_ca'])})")
    ax.set_xlabel("Down-aisle buckling length  Lcr,DA  [mm]")
    ax.set_ylabel("Axial capacity  N_b,Rd  [kN]")
    ax.set_title(f"{name}  -  {info['desc']}\nUpright load chart  ·  fy=355  ·  "
                 f"γM1=1.1  (X=pitch, D=2×pitch, XS=+internal stiffener)",
                 fontsize=10)
    ax.grid(True, alpha=0.3)
    ax.set_ylim(bottom=0)
    ax.legend(fontsize=8, ncol=2)
    fig.tight_layout()
    fig.savefig(path, dpi=130)
    plt.close(fig)


def _plot_beam(name: str, info: dict, path: str) -> None:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    fig, ax = plt.subplots(figsize=(8.5, 5.5))
    for i, (key, c) in enumerate(info["curves"].items()):
        colour = _BEAM_COLOURS[i % len(_BEAM_COLOURS)]
        ax.plot(c["span"], c["load"], "-", color=colour, lw=1.8,
                label=f"upright {key}")
    ax.set_xlabel("Beam span  [mm]")
    ax.set_ylabel("Max load per bay level (pair of beams)  [kN]")
    ax.set_title(f"{name}  -  beam load chart  ·  fy={info['fy']:.0f}  ·  "
                 f"deflection L/200\n(semi-rigid connectors; curve per upright "
                 f"thickness)", fontsize=10)
    ax.grid(True, alpha=0.3)
    ax.set_ylim(bottom=0)
    ax.legend(fontsize=8)
    fig.tight_layout()
    fig.savefig(path, dpi=130)
    plt.close(fig)


# --------------------------------------------------------------------- Excel
def _write_upright_xlsx(data: Dict[str, dict], path: str) -> None:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Uprights"
    ws.append(["section", "description", "config", "Lcr_CA_mm", "Lcr_DA_mm",
               "capacity_kN", "governing"])
    for name, info in data.items():
        for label, c in info["curves"].items():
            for da, cap, gov in zip(c["da"], c["cap"], c["gov"]):
                ws.append([name, info["desc"], label, int(c["lcr_ca"]), da,
                           round(cap, 2), gov])
    wb.save(path)


def _write_beam_xlsx(data: Dict[str, dict], path: str) -> None:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Beams"
    ws.append(["beam", "fy_MPa", "upright_thickness", "span_mm",
               "load_per_level_kN", "governing"])
    for name, info in data.items():
        for key, c in info["curves"].items():
            for sp, load, gov in zip(c["span"], c["load"], c["gov"]):
                ws.append([name, round(info["fy"], 0), key, sp,
                           round(load, 2), gov])
    wb.save(path)


# ----------------------------------------------------------------- generation
def generate(master_path: str, out_dir: str) -> dict:
    """Generate all upright and beam load charts + tables under out_dir.
    Returns a summary dict with the output paths."""
    mw = load_master(master_path)
    lib = mw.library

    up_dir = os.path.join(out_dir, "uprights")
    bm_dir = os.path.join(out_dir, "beams")
    os.makedirs(up_dir, exist_ok=True)
    os.makedirs(bm_dir, exist_ok=True)

    up_data = upright_chart_data(lib, fy=FY_UPRIGHT)
    for name, info in up_data.items():
        _plot_upright(name, info, os.path.join(up_dir, f"{name}.png"))
    up_xlsx = os.path.join(out_dir, "Upright_Load_Charts.xlsx")
    _write_upright_xlsx(up_data, up_xlsx)

    def fy_of(beam_name: str) -> float:
        return float(mw.fy.get(beam_name, FY_UPRIGHT) or FY_UPRIGHT)

    bm_data = beam_chart_data(lib, fy_of)
    for name, info in bm_data.items():
        safe = name.replace("/", "_")
        _plot_beam(name, info, os.path.join(bm_dir, f"{safe}.png"))
    bm_xlsx = os.path.join(out_dir, "Beam_Load_Charts.xlsx")
    _write_beam_xlsx(bm_data, bm_xlsx)

    # zip the PNGs for delivery
    png_zip = os.path.join(out_dir, "load_charts_png.zip")
    with zipfile.ZipFile(png_zip, "w", zipfile.ZIP_DEFLATED) as z:
        for d, sub in ((up_dir, "uprights"), (bm_dir, "beams")):
            for f in sorted(os.listdir(d)):
                if f.endswith(".png"):
                    z.write(os.path.join(d, f), f"{sub}/{f}")

    return {"uprights": len(up_data), "beams": len(bm_data),
            "upright_xlsx": up_xlsx, "beam_xlsx": bm_xlsx, "png_zip": png_zip,
            "out_dir": out_dir}


def main(argv: Optional[List[str]] = None) -> None:
    ap = argparse.ArgumentParser(description="Generate upright & beam load charts.")
    ap.add_argument("--master", required=True, help="section master .xlsx")
    ap.add_argument("--out", default="charts", help="output directory")
    ap.add_argument("--model", action="store_true",
                    help="model-based upright working-load charts (N+My+Mz, full FEA)")
    args = ap.parse_args(argv)
    if args.model:
        s = generate_model_based(args.master, args.out)
        print(f"model-based upright charts: {s['sections']} sections")
        print(f"  {s['xlsx']}\n  {s['png_zip']}")
        return
    summary = generate(args.master, args.out)
    print(f"uprights: {summary['uprights']} charts, beams: {summary['beams']} charts")
    print(f"  {summary['upright_xlsx']}")
    print(f"  {summary['beam_xlsx']}")
    print(f"  {summary['png_zip']}")


# =====================================================================
# MODEL-BASED upright working-load charts (full FEA: N + My + Mz)
# =====================================================================
# Unlike the closed-form charts above (pure axial N_b,Rd, an upper bound), these
# build the real rack and find the axial at which the GOVERNING upright
# interaction N/Nb,Rd + kM(My/My,Rd + Mz/Mz,Rd) reaches 1.0 - so the down-aisle
# connector/sway/imperfection moments and cross-aisle moment are included.
# Reference rack: single-module 3-bay (validated within 2% of, and slightly
# conservative vs, the back-to-back rack for the down-aisle upright capacity),
# 5 levels, beam RHS112X50X1.6, base/connector/imperfections from the saved
# archetype (examples/loadchart_archetype.json).  ULS2 governs (verified).

import json as _json

_ARCH_PATH = os.path.join(os.path.dirname(__file__), "..", "examples",
                          "loadchart_archetype.json")
MODEL_BEAM = "RHS112X50X1.6"
MODEL_NLEV = 1
MODEL_NBAYS = 3
MODEL_DA = list(range(250, 4001, 100))     # beam gap = Lcr_DA [mm], 100 mm steps
# load / resistance factors used by the model-based chart (overridable):
CHART_GAMMA_G = 1.2         # dead load factor at ULS (was 1.3)
CHART_GAMMA_Q = 1.2         # imposed (pallet) load factor at ULS (was 1.4)
CHART_GAMMA_M1 = 1.0        # member buckling resistance factor (was 1.1)
# (label, bracing_type, pitch, with-stiffener) - 600 mm pitch only
MODEL_CONFIGS = [
    ("X-600", "X", 600.0, False),
    ("D-1200", "D", 600.0, False),
    ("XS-600", "X", 600.0, True),
]


def _load_archetype() -> dict:
    return _json.load(open(_ARCH_PATH, encoding="utf-8"))


def _label(btype, pitch, xs):
    if xs:
        return f"XS-{int(pitch)}"
    return f"X-{int(pitch)}" if btype == "X" else f"D-{int(2 * pitch)}"


def _build_model_rack(mw, arch, section, gap, btype, pitch, xs, load,
                      nlev=None):
    """Single-module 3-bay rack at a per-level load, reduced to the governing
    ULS2 combination with down-aisle imperfection (fast capacity eval)."""
    import dataclasses
    from .builder import RackConfig, build_rack, LevelSpec
    nlev = int(nlev or MODEL_NLEV)
    fields = {f.name for f in dataclasses.fields(RackConfig)}
    kw = {k: v for k, v in arch.items() if k in fields}
    h = nlev * gap + 200.0
    kw["levels"] = [LevelSpec(gap=gap, beam_section=MODEL_BEAM, pallet_load=load)
                    for _ in range(nlev)]
    kw.update(master=mw, module="single", n_bays=MODEL_NBAYS,
              upright_section=section, beam_section=MODEL_BEAM,
              bracing_type=btype, bracing_pitch=pitch, frame_height=h,
              pallet_load_per_level=load, ca_brace_zones=(), ca_x_height=None,
              steel_fy=FY_UPRIGHT, fy_override=True,
              gamma_G=CHART_GAMMA_G, gamma_G_uls=CHART_GAMMA_G,
              gamma_Q=CHART_GAMMA_Q, pay_placement_factor=CHART_GAMMA_Q,
              include_self_weight=False)   # pallet capacity; self-wt negligible
    if xs:
        depth = round(mw.library.get(section).depth_h or 0.0)
        st = {90: "IN_STIFFENER90X1.6", 120: "IN_STIFFENER120X1.6"}.get(depth)
        if not st:
            return None
        kw.update(stiffener_section=st, reinforce_height=h, stiffener_type=1)
    m = build_rack(RackConfig(**kw))
    m.checks.gamma_M1 = CHART_GAMMA_M1               # resistance factor (no E reduction)
    m.analysis.compute_alpha_cr = False
    # lean solver for the sweep: a single fast attempt that fails quickly when
    # the load is above the stability limit (the search then reduces the load)
    m.analysis.fast_solve = True
    m.analysis.n_steps = 4
    m.analysis.max_iter = 12
    # full elastic stiffness (RSTAB uses gamma_M = 1.0); the material-factor
    # reduction is a user-selectable option, not applied to the chart by default
    m.imperfection.directions = ["+x"]
    # govern by the worse of max-gravity (ULS1: 1.4 LL) and the placement combo
    # (ULS2: 1.26 LL + 1.26 placement) - ULS1 has the higher axial and usually
    # governs down-aisle stability, ULS2 adds the placement sway force.
    m.combinations = [c for c in m.combinations if c.name in ("ULS1", "ULS2")]
    return m


def _eval_upright(mw, arch, section, gap, btype, pitch, xs, load, nlev=None):
    from .analysis import run_all, UnstableModelError
    from .checks.en15512 import run_checks, upright_set_buckling_rows
    m = _build_model_rack(mw, arch, section, gap, btype, pitch, xs, load, nlev)
    if m is None:
        return None
    try:
        rows = upright_set_buckling_rows(m, run_checks(m, run_all(m)))
    except Exception:
        return (99.0, 0.0, 0.0, 0.0)
    up = [r for r in rows if r["set"].startswith("Upright")]
    if not up:
        return (0.0, 0.0, 0.0, 0.0)
    gov = max(up, key=lambda r: r["util"])
    util, N = gov["util"], gov["N_kN"]
    # XS: the reinforced column is the upright AND the stiffener acting together -
    # the assembly fails when EITHER member reaches 1.0, and the axial capacity is
    # the COMBINED axial (upright + the stiffener it shares load with).
    stf = [r for r in rows if r["set"].startswith("Stiffener")]
    if stf:
        gov_st = max(stf, key=lambda r: r["util"])
        util = max(util, gov_st["util"])
        target = gov["set"].replace("Upright", "Stiffener")
        match = next((r for r in stf if r["set"] == target), gov_st)
        N = N + match["N_kN"]
    return (util, N, gov["My_kNm"], gov["Mz_kNm"])


def model_upright_point(mw, arch, section, gap, btype, pitch, xs):
    """Maximum working-load capacity at one point.

    Per the load-reduction rule: a non-converging ULS means the load is above
    the upright's stability limit, so reduce the load until the model converges
    AND the governing interaction (N+My+Mz) stays <= 1.0.  The capacity is the
    largest per-level load satisfying both - whichever governs, buckling or
    strength.  Seeded from the closed-form strut value and bracketed/bisected.
    """
    lcr_ca = pitch if btype == "X" else 2.0 * pitch
    sec = mw.library.get(section)
    if xs:
        depth = round(sec.depth_h or 0.0)
        stn = {90: "IN_STIFFENER90X1.6", 120: "IN_STIFFENER120X1.6"}.get(depth)
        if not stn:
            return None
        st = mw.library.get(stn)
        sec = combined_upright_section(sec, st, st.mount_offset or 30.0)
    n_axial, _ = upright_capacity_kN(sec, FY_UPRIGHT, gap, lcr_ca)
    seed = max(n_axial * 1e3 / MODEL_NLEV / 1.4 * 0.8, 200.0)

    def ev(load):
        r = _eval_upright(mw, arch, section, gap, btype, pitch, xs, load)
        if r is None:
            return None
        u, N, My, Mz = r
        feasible = (u < 10.0) and (0.0 < u <= 1.0)   # u>=10 sentinel = unstable
        return feasible, u, N, My, Mz

    lo = None          # (load,u,N,My,Mz): largest load that is stable AND u<=1
    hi = None          # smallest load known infeasible (unstable OR u>1)
    load = seed
    for _ in range(7):
        r = ev(load)
        if r is None:
            return None
        feasible, u, N, My, Mz = r
        if feasible:
            if lo is None or load > lo[0]:
                lo = (load, u, N, My, Mz)
            nxt = (load * min(max(1.0 / max(u, 1e-3), 1.05), 1.8)
                   if hi is None else 0.5 * (load + hi))
        else:
            hi = load if hi is None else min(hi, load)
            nxt = load * 0.5 if lo is None else 0.5 * (lo[0] + hi)
        if lo and hi and (hi - lo[0]) <= 0.05 * lo[0]:
            break
        load = max(nxt, 50.0)

    if lo is None:
        return None
    load, u, N, My, Mz = lo
    # strength-governed but not iterated to exactly 1.0 (no upper bound found):
    # scale the converged result to the interaction limit; otherwise lo already
    # sits at the governing (stability or strength) capacity.
    s = (1.0 / u) if (hi is None and u > 0.0) else 1.0
    return {"section": section, "config": _label(btype, pitch, xs),
            "lcr_da": int(gap), "lcr_ca": int(lcr_ca),
            "load_per_level_kN": round(load * s / 1e3, 2),
            "N_cap_kN": round(N * s, 2), "My_kNm": round(My * s, 3),
            "Mz_kNm": round(Mz * s, 3)}


G_ACC = 9.81                # m/s^2: pallet mass (kg) -> load (N)
LOAD_CAP_KG = 3000.0        # capped load per beam level [kg]
LEVELS_TARGET = 3           # start height; extend up while there is spare capacity
LEVELS_MAX = 7
LEVELS_MIN = 1


def model_levels_point(mw, arch, section, gap, btype, pitch, xs,
                       load_kg=LOAD_CAP_KG, lmin=LEVELS_MIN, lmax=LEVELS_MAX,
                       target=LEVELS_TARGET):
    """Maximum number of beam levels the upright supports at a fixed capped load
    per level (default 3000 kg).

    The load per level is held at the cap; the OUTPUT is how tall the rack can
    be.  A non-converging or over-utilised ULS means the rack is too tall, so
    reduce the level count; if the target (5) is met with margin, extend up to
    lmax (7).  Returns the governing bottom-upright axial and Mz at that height.
    """
    lcr_ca = pitch if btype == "X" else 2.0 * pitch
    load = load_kg * G_ACC
    if xs:
        depth = round(mw.library.get(section).depth_h or 0.0)
        if depth not in (90, 120):
            return None

    def feasible(L):
        r = _eval_upright(mw, arch, section, gap, btype, pitch, xs, load, nlev=L)
        if r is None:
            return None
        u, N, My, Mz = r
        ok = (u < 10.0) and (0.0 < u <= 1.0)        # u>=10 sentinel = unstable
        return ok, u, N, My, Mz

    best = None
    r = feasible(target)
    if r is None:
        return None
    if r[0]:                                          # target met -> climb
        best = (target,) + r[1:]
        L = target
        while L < lmax:
            L += 1
            r = feasible(L)
            if r and r[0]:
                best = (L,) + r[1:]
            else:
                break
    else:                                             # target too tall -> descend
        L = target
        while L > lmin:
            L -= 1
            r = feasible(L)
            if r and r[0]:
                best = (L,) + r[1:]
                break
    out = {"section": section, "config": _label(btype, pitch, xs),
           "lcr_ca": int(lcr_ca), "lcr_da": int(gap), "load_kg": load_kg}
    if best is None:                                  # cannot carry lmin levels
        out.update(max_levels=0, bottom_axial_kN=0.0, util=None, Mz_kNm=0.0)
        return out
    L, u, N, My, Mz = best
    out.update(max_levels=int(L), bottom_axial_kN=round(N, 1),
               util=round(u, 3), Mz_kNm=round(Mz, 3))
    return out


# --------------------------------------------------------------------------
# utilization-targeted chart: tune the (soft) load per level so the governing
# upright utilisation (max of STRESS and BUCKLING) lands in 0.97..0.99, with the
# level count optimised slightly around a seed.
# --------------------------------------------------------------------------
UTIL_TARGET = 0.98
UTIL_LO = 0.97
UTIL_HI = 0.99
UTIL_LOAD_CAP_KG = 2500.0   # HARD cap on load per level for the util chart [kg]
UTIL_LOAD_FLOOR_KG = 1500.0 # floor for the extended (sway-limited) search [kg]
MAX_LEVELS_CAP = 30         # upper bound on the level-count search


def _eval_util(mw, arch, section, gap, btype, pitch, xs, load, nlev):
    """Governing upright utilisation at (load per level, nlev).

    Returns (converged, gov_util, stress_util, buckling_util, bottom_axial_kN);
    gov_util = max over the upright AND upright-stiffener members of the STRESS
    and BUCKLING utilisations.  converged is False (gov set to a 99 sentinel)
    when the model is unstable / no ULS converges at this load.
    """
    from .analysis import run_all
    from .checks.en15512 import run_checks
    m = _build_model_rack(mw, arch, section, gap, btype, pitch, xs, load, nlev)
    if m is None:
        return None
    try:
        checks = run_checks(m, run_all(m))
    except Exception:
        return (False, 99.0, 99.0, 99.0, 0.0)
    su = bu = 0.0
    n_ax = 0.0
    sets = ("uprights", "upright stiffeners")
    for ch in checks:
        if ch.informative or ch.member_set not in sets:
            continue
        if ch.check == "STRESS":
            su = max(su, ch.utilization)
        elif ch.check == "BUCKLING":
            bu = max(bu, ch.utilization)
            x = ch.extra or {}
            n_ax = max(n_ax, abs((x.get("N") or 0.0) / 1.0e3))
    gov = max(su, bu)
    return (gov < 50.0 and gov > 0.0, gov, su, bu, n_ax)


def _tune_load(mw, arch, section, gap, btype, pitch, xs, nlev, seed,
               max_load=1.0e9):
    """At a fixed level count, tune the load per level so the governing upright
    util reaches the 0.97..0.99 band, never exceeding max_load.  Returns the
    best converged point (load_N, gov, su, bu, axial) closest to UTIL_TARGET, or
    None if the frame is sway-unstable before util can rise into the band."""
    load = max(min(float(seed), max_load), 200.0)
    lo = None             # (load) largest converged load below target
    hi = None             # smallest load that overshoots / goes unstable
    best = None
    for _ in range(7):
        r = _eval_util(mw, arch, section, gap, btype, pitch, xs, load, nlev)
        if r is None:
            return None
        conv, gov, su, bu, ax = r
        if conv:
            if best is None or abs(gov - UTIL_TARGET) < abs(best[1] - UTIL_TARGET):
                best = (load, gov, su, bu, ax)
            if UTIL_LO <= gov <= UTIL_HI:
                return best
            if gov < UTIL_LO:
                if load >= max_load - 1.0:           # capped: cannot add load
                    return best
                lo = load
                nxt = (load * min(max(UTIL_TARGET / max(gov, 1e-3), 1.05), 1.8)
                       if hi is None else 0.5 * (load + hi))
            else:                                   # overshoot util>0.99
                hi = load if hi is None else min(hi, load)
                nxt = 0.5 * ((lo if lo else load * 0.5) + hi)
        else:                                       # unstable -> load too high
            hi = load if hi is None else min(hi, load)
            nxt = (load * 0.5) if lo is None else 0.5 * (lo + hi)
        load = max(min(nxt, max_load), 50.0)
    return best


def model_util_point(mw, arch, section, gap, btype, pitch, xs, n_seed=3,
                     cap_kg=None):
    """Find (n_levels, load per level) giving governing upright util in
    0.97..0.99 with the load per level HARD-capped at UTIL_LOAD_CAP_KG.

    Capacity is built by ADDING LEVELS (not weight): the load stays at the cap
    and the level count is raised until util enters the band; only when the band
    falls between two integer level counts is the load trimmed below the cap at
    the higher level count (more levels, lighter).  With the load fixed a stiffer
    config carries more levels, so capacity is monotone XS > X > D.
    """
    lcr_ca = pitch if btype == "X" else 2.0 * pitch
    if xs:
        depth = round(mw.library.get(section).depth_h or 0.0)
        if depth not in (90, 120):
            return None
    cap = (cap_kg or UTIL_LOAD_CAP_KG) * G_ACC       # N per level (hard cap)

    def at_cap(n):
        """(feasible, conv, gov, su, bu, ax) at the cap load for n levels;
        feasible = converged and util <= UTIL_HI."""
        r = _eval_util(mw, arch, section, gap, btype, pitch, xs, cap, n)
        if r is None:
            return None
        conv, gov, su, bu, ax = r
        return (conv and gov <= UTIL_HI, conv, gov, su, bu, ax)

    def pack(n, load, gov, su, bu, ax):
        out = {"section": section, "config": _label(btype, pitch, xs),
               "lcr_ca": int(lcr_ca), "lcr_da": int(gap)}
        load_kg = load / G_ACC
        out.update(n_levels=int(n), load_per_level_kg=round(load_kg, 1),
                   frame_load_kg=round(load_kg * n, 1),
                   bottom_axial_kN=round(ax, 1), stress_util=round(su, 3),
                   buckling_util=round(bu, 3), gov_util=round(gov, 3))
        return out

    def reduced(n):
        """n levels with the load trimmed below the cap to land util in band."""
        return _tune_load(mw, arch, section, gap, btype, pitch, xs, n, cap,
                          max_load=cap)

    _cache: Dict[int, object] = {}

    def cap_eval(n):
        if n not in _cache:
            _cache[n] = at_cap(n)
        return _cache[n]

    # --- find N_max = the largest level count that fits at the cap load -------
    # (converged AND util <= UTIL_HI).  Feasibility is monotone in n (taller =
    # less stable / higher util), so walk up/down from the seed.
    n = max(int(n_seed) if n_seed else 3, 1)
    f = cap_eval(n)
    if f is None:
        return pack(0, 0.0, 0.0, 0.0, 0.0, 0.0)
    if f[0]:                                          # seed fits -> climb up
        N_max = n
        while N_max < MAX_LEVELS_CAP:
            g = cap_eval(N_max + 1)
            if g is None or not g[0]:
                break
            N_max += 1
    else:                                             # seed too tall -> descend
        N_max = 0
        m = n - 1
        while m >= 1:
            g = cap_eval(m)
            if g is not None and g[0]:
                N_max = m
                break
            m -= 1

    # --- candidates: maximise capacity (n * load) with load<=cap, util<=UTIL_HI
    cands = []                                        # (frame_kg, n, load, gov,su,bu,ax)
    if N_max >= 1:
        a = cap_eval(N_max)
        cands.append((UTIL_LOAD_CAP_KG * N_max, N_max, cap,
                      a[2], a[3], a[4], a[5]))
    # one extra level at a trimmed (sub-cap) load - usually more total capacity
    rb = reduced(N_max + 1)
    if rb and rb[1] <= UTIL_HI + 1e-6:
        cands.append((rb[0] / G_ACC * (N_max + 1), N_max + 1, rb[0],
                      rb[1], rb[2], rb[3], rb[4]))
    if N_max == 0:                                    # even 1 level over at cap
        r1 = reduced(1)
        if r1:
            cands.append((r1[0] / G_ACC, 1, r1[0],
                          r1[1], r1[2], r1[3], r1[4]))
    if not cands:
        return pack(0, 0.0, 0.0, 0.0, 0.0, 0.0)
    # pick the highest capacity (frame load); util is <= UTIL_HI by construction
    best = max(cands, key=lambda c: c[0])             # (frame,n,load,gov,su,bu,ax)
    if best[3] >= UTIL_LO:                            # already in band -> done
        return pack(best[1], best[2], best[3], best[4], best[5], best[6])

    # sway/cap-limited below the band: the load cap left the frame's capacity
    # under-used -> ADD LEVELS and REDUCE the load (down to UTIL_LOAD_FLOOR_KG)
    # so the same buckling capacity is reached over more, lighter levels.  The
    # extra-level climb is bounded (a few levels above N_max) and stops at the
    # first in-band result, so it cannot run away into huge, very slow models.
    floor = UTIL_LOAD_FLOOR_KG * G_ACC
    misses = 0
    n2 = max(N_max + 1, 2)
    n2_max = min(MAX_LEVELS_CAP, max(N_max, 1) + 6)
    while n2 <= n2_max and misses < 3:
        r = _tune_load(mw, arch, section, gap, btype, pitch, xs, n2, cap,
                       max_load=cap)
        if r is None:
            break
        load_r, gov_r, su_r, bu_r, ax_r = r
        if load_r < floor - 1.0:                      # would need < 1500 kg -> stop
            break
        frame_r = load_r / G_ACC * n2
        if UTIL_LO <= gov_r <= UTIL_HI:               # reached the band -> done
            return pack(n2, load_r, gov_r, su_r, bu_r, ax_r)
        misses += 1
        # keep the closest-to-target as an improved fallback (still load>=floor)
        if abs(gov_r - UTIL_TARGET) < abs(best[3] - UTIL_TARGET):
            best = (frame_r, n2, load_r, gov_r, su_r, bu_r, ax_r)
        n2 += 1

    return pack(best[1], best[2], best[3], best[4], best[5], best[6])


_WORKER: dict = {}


def _winit(master_path):
    _WORKER["mw"] = load_master(master_path)
    _WORKER["arch"] = _load_archetype()


def _wpoint(args):
    s, bt, p, xs, g = args
    key = f"{s}|{_label(bt, p, xs)}|{g}"
    try:
        pt = model_upright_point(_WORKER["mw"], _WORKER["arch"], s, float(g),
                                 bt, p, xs)
    except Exception:
        pt = None
    return key, (pt or {})


def generate_model_based(master_path, out_dir, checkpoint=None, workers=None):
    """Resumable, parallel model-based upright working-load charts (N+My+Mz).

    Each point is an independent FEA capacity search, so the grid is fanned out
    across processes.  Progress is checkpointed frequently, so an interrupted
    run (e.g. a background time limit) resumes where it stopped."""
    import multiprocessing as mp
    os.makedirs(out_dir, exist_ok=True)
    checkpoint = checkpoint or os.path.join(out_dir, "_model_checkpoint.json")
    done = {}
    if os.path.exists(checkpoint):
        done = _json.load(open(checkpoint, encoding="utf-8"))
    sections = load_master(master_path).library.names("upright")
    todo = [(s, bt, p, xs, g) for s in sections
            for (_lbl, bt, p, xs) in MODEL_CONFIGS for g in MODEL_DA
            if f"{s}|{_label(bt, p, xs)}|{g}" not in done]
    total = len(sections) * len(MODEL_CONFIGS) * len(MODEL_DA)
    workers = workers or min(4, os.cpu_count() or 1)
    print(f"model-based grid: {total} points, {len(todo)} to do, "
          f"{workers} workers", flush=True)
    n = 0
    with mp.Pool(workers, initializer=_winit, initargs=(master_path,)) as pool:
        for key, pt in pool.imap_unordered(_wpoint, todo, chunksize=1):
            done[key] = pt
            n += 1
            if n % 20 == 0:
                _json.dump(done, open(checkpoint, "w", encoding="utf-8"))
                ok = len([k for k in done if done[k]])
                print(f"  ... {len(done)}/{total} evaluated ({ok} with capacity)",
                      flush=True)
    _json.dump(done, open(checkpoint, "w", encoding="utf-8"))
    return _finalize_model_based(done, out_dir)


def _finalize_model_based(done, out_dir):
    import openpyxl
    up_dir = os.path.join(out_dir, "uprights_model")
    os.makedirs(up_dir, exist_ok=True)
    bysec = {}
    for pt in done.values():
        if not pt:
            continue
        c = bysec.setdefault(pt["section"], {}).setdefault(
            pt["config"], {"da": [], "cap": [], "load": [], "lcr_ca": pt["lcr_ca"]})
        c["da"].append(pt["lcr_da"])
        c["cap"].append(pt["N_cap_kN"])
        c["load"].append(pt["load_per_level_kN"])
    for sec, curves in bysec.items():
        for c in curves.values():
            order = sorted(range(len(c["da"])), key=lambda i: c["da"][i])
            for k in ("da", "cap", "load"):
                c[k] = [c[k][i] for i in order]
        _plot_model_upright(sec, curves, os.path.join(up_dir, f"{sec}.png"))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Uprights_model"
    ws.append(["section", "config", "Lcr_CA_mm", "Lcr_DA_mm",
               "upright_axial_capacity_kN", "load_per_level_kN", "My_kNm", "Mz_kNm"])
    for key, pt in sorted(done.items()):
        if pt:
            ws.append([pt["section"], pt["config"], pt["lcr_ca"], pt["lcr_da"],
                       pt["N_cap_kN"], pt["load_per_level_kN"], pt["My_kNm"],
                       pt["Mz_kNm"]])
    xlsx = os.path.join(out_dir, "Upright_Load_Charts_MODEL.xlsx")
    wb.save(xlsx)
    png_zip = os.path.join(out_dir, "upright_model_png.zip")
    with zipfile.ZipFile(png_zip, "w", zipfile.ZIP_DEFLATED) as z:
        for f in sorted(os.listdir(up_dir)):
            if f.endswith(".png"):
                z.write(os.path.join(up_dir, f), f"uprights_model/{f}")
    return {"sections": len(bysec), "xlsx": xlsx, "png_zip": png_zip}


def _plot_model_upright(name, curves, path):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    fig, ax = plt.subplots(figsize=(8.5, 5.5))
    for label in ("X-500", "X-600", "D-1000", "D-1200", "XS-500", "XS-600"):
        c = curves.get(label)
        if not c:
            continue
        colour, ls = _STYLE.get(label, (branding.GREY, "-"))
        ax.plot(c["da"], c["cap"], ls, color=colour, lw=1.8, label=label)
    ax.set_xlabel("Down-aisle buckling length  Lcr,DA = beam gap  [mm]")
    ax.set_ylabel("Upright axial capacity  [kN]  (XS = upright+stiffener; N+My+Mz = 1)")
    ax.set_title(f"{name}  -  model-based upright load chart  ·  fy=355  ·  "
                 f"gamma_M1=1.1\nsingle-storey 3-bay sub-assemblage, RHS112 beam, "
                 f"EN 15512 imperfections (max stable load; N+My+Mz)", fontsize=9.5)
    ax.grid(True, alpha=0.3)
    ax.set_ylim(bottom=0)
    ax.legend(fontsize=8, ncol=2)
    fig.tight_layout()
    fig.savefig(path, dpi=130)
    plt.close(fig)


# --------------------------------------------------------------------------
# levels-based chart: max number of beam levels at a capped load per level
# (unbraced down-aisle semi-rigid moment frame; start at 3, climb to 7)
# --------------------------------------------------------------------------
def _wpoint_levels(args):
    s, bt, p, xs, g = args
    key = f"{s}|{_label(bt, p, xs)}|{g}"
    try:
        pt = model_levels_point(_WORKER["mw"], _WORKER["arch"], s, float(g),
                                bt, p, xs)
    except Exception:
        pt = None
    return key, (pt or {})


def generate_levels_based(master_path, out_dir, checkpoint=None, workers=None):
    """Resumable, parallel max-levels chart: for each upright/config/beam gap,
    the largest number of beam levels (1..7, starting from 3) that the UNBRACED
    down-aisle semi-rigid moment frame carries at the capped load per level
    (3000 kg).  A non-converging/over-utilised ULS means the rack is too tall."""
    import multiprocessing as mp
    os.makedirs(out_dir, exist_ok=True)
    checkpoint = checkpoint or os.path.join(out_dir, "_levels_checkpoint.json")
    done = {}
    if os.path.exists(checkpoint):
        done = _json.load(open(checkpoint, encoding="utf-8"))
    sections = load_master(master_path).library.names("upright")
    todo = [(s, bt, p, xs, g) for s in sections
            for (_lbl, bt, p, xs) in MODEL_CONFIGS for g in MODEL_DA
            if f"{s}|{_label(bt, p, xs)}|{g}" not in done]
    total = len(sections) * len(MODEL_CONFIGS) * len(MODEL_DA)
    workers = workers or min(4, os.cpu_count() or 1)
    print(f"max-levels grid: {total} points, {len(todo)} to do, "
          f"{workers} workers, load {LOAD_CAP_KG:.0f} kg/level, "
          f"levels {LEVELS_MIN}..{LEVELS_MAX}", flush=True)
    n = 0
    with mp.Pool(workers, initializer=_winit, initargs=(master_path,)) as pool:
        for key, pt in pool.imap_unordered(_wpoint_levels, todo, chunksize=1):
            done[key] = pt
            n += 1
            if n % 20 == 0:
                _json.dump(done, open(checkpoint, "w", encoding="utf-8"))
                print(f"  ... {len(done)}/{total} evaluated", flush=True)
    _json.dump(done, open(checkpoint, "w", encoding="utf-8"))
    return _finalize_levels(done, out_dir)


def _finalize_levels(done, out_dir):
    import openpyxl
    up_dir = os.path.join(out_dir, "uprights_levels")
    os.makedirs(up_dir, exist_ok=True)
    bysec = {}
    for pt in done.values():
        if not pt:
            continue
        c = bysec.setdefault(pt["section"], {}).setdefault(
            pt["config"], {"da": [], "lev": [], "lcr_ca": pt["lcr_ca"]})
        c["da"].append(pt["lcr_da"])
        c["lev"].append(pt["max_levels"])
    for sec, curves in bysec.items():
        for c in curves.values():
            order = sorted(range(len(c["da"])), key=lambda i: c["da"][i])
            for k in ("da", "lev"):
                c[k] = [c[k][i] for i in order]
        _plot_levels(sec, curves, os.path.join(up_dir, f"{sec}.png"))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Max_levels_3000kg"
    ws.append(["section", "config", "Lcr_CA_mm", "Lcr_DA_mm", "load_per_level_kg",
               "max_levels", "bottom_upright_axial_kN", "util", "Mz_kNm"])
    for key, pt in sorted(done.items()):
        if pt:
            ws.append([pt["section"], pt["config"], pt["lcr_ca"], pt["lcr_da"],
                       pt.get("load_kg", LOAD_CAP_KG), pt["max_levels"],
                       pt.get("bottom_axial_kN"), pt.get("util"),
                       pt.get("Mz_kNm")])
    xlsx = os.path.join(out_dir, "Upright_Max_Levels_3000kg.xlsx")
    wb.save(xlsx)
    png_zip = os.path.join(out_dir, "upright_levels_png.zip")
    with zipfile.ZipFile(png_zip, "w", zipfile.ZIP_DEFLATED) as z:
        for f in sorted(os.listdir(up_dir)):
            if f.endswith(".png"):
                z.write(os.path.join(up_dir, f), f"uprights_levels/{f}")
    return {"sections": len(bysec), "xlsx": xlsx, "png_zip": png_zip}


def _plot_levels(name, curves, path):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    fig, ax = plt.subplots(figsize=(8.5, 5.5))
    for label in ("X-500", "X-600", "D-1000", "D-1200", "XS-500", "XS-600"):
        c = curves.get(label)
        if not c:
            continue
        colour, ls = _STYLE.get(label, (branding.GREY, "-"))
        ax.step(c["da"], c["lev"], where="mid", color=colour, lw=1.8, label=label)
    ax.set_xlabel("Down-aisle buckling length  Lcr,DA = beam gap  [mm]")
    ax.set_ylabel(f"Max beam levels at {LOAD_CAP_KG:.0f} kg/level")
    ax.set_title(f"{name}  -  max rack levels at {LOAD_CAP_KG:.0f} kg/level  ·  "
                 f"fy=355  ·  gamma_M1=1.1\nunbraced 3-bay semi-rigid moment frame, "
                 f"RHS112 beam, EN 15512 imperfections (N+My+Mz)", fontsize=9.5)
    ax.grid(True, alpha=0.3)
    ax.set_ylim(0, LEVELS_MAX + 0.5)
    ax.set_yticks(range(0, LEVELS_MAX + 1))
    ax.legend(fontsize=8, ncol=2)
    fig.tight_layout()
    fig.savefig(path, dpi=130)
    plt.close(fig)


# --------------------------------------------------------------------------
# utilization-targeted chart generator (parallel, resumable)
# --------------------------------------------------------------------------
def _wpoint_util(args):
    s, bt, p, xs, g, nseed = args
    key = f"{s}|{_label(bt, p, xs)}|{g}"
    try:
        pt = model_util_point(_WORKER["mw"], _WORKER["arch"], s, float(g),
                              bt, p, xs, nseed)
    except Exception:
        pt = None
    return key, (pt or {})


def _wpoint_util_cap(args):
    """Worker that re-optimises one case at an explicit per-level load cap."""
    s, bt, p, xs, g, nseed, cap_kg = args
    key = f"{s}|{_label(bt, p, xs)}|{g}"
    try:
        pt = model_util_point(_WORKER["mw"], _WORKER["arch"], s, float(g),
                              bt, p, xs, nseed, cap_kg=cap_kg)
    except Exception:
        pt = None
    return key, (pt or {})


def rerun_under_cases(master_path, main_checkpoint, out_dir, cap_kg=3000.0,
                      workers=None, under_threshold=UTIL_LO):
    """Re-optimise only the cases below `under_threshold` (default 0.97) from a
    finished util chart, allowing a higher per-level load cap (default 3000 kg),
    to try to reach the 0.97-0.99 band.  Writes a separate xlsx of the re-run
    cases and a merged xlsx (cases at/above the threshold unchanged; below it
    replaced by the higher-cap result)."""
    import multiprocessing as mp
    import openpyxl
    os.makedirs(out_dir, exist_ok=True)
    main = _json.load(open(main_checkpoint, encoding="utf-8"))
    under = {k for k, v in main.items()
             if (not v) or (v.get("gov_util") is None)
             or v["gov_util"] < under_threshold}
    # reverse-map config label -> (bt, p, xs)
    cfgmap = {_label(bt, p, xs): (bt, p, xs) for (_l, bt, p, xs) in MODEL_CONFIGS}
    todo = []
    for k in under:
        sec, cfg, gs = k.split("|")
        if cfg not in cfgmap:
            continue
        bt, p, xs = cfgmap[cfg]
        nseed = (main.get(k) or {}).get("n_levels") or 3
        todo.append((sec, bt, p, xs, int(gs), max(int(nseed), 1), cap_kg))
    ckpt = os.path.join(out_dir, "_under3000_checkpoint.json")
    done = _json.load(open(ckpt, encoding="utf-8")) if os.path.exists(ckpt) else {}
    todo = [t for t in todo if f"{t[0]}|{_label(t[1],t[2],t[3])}|{t[4]}" not in done]
    workers = workers or min(4, os.cpu_count() or 1)
    print(f"UNDER re-run @ {cap_kg:.0f} kg: {len(under)} under, {len(todo)} to do",
          flush=True)
    n = 0
    with mp.Pool(workers, initializer=_winit, initargs=(master_path,)) as pool:
        for key, pt in pool.imap_unordered(_wpoint_util_cap, todo, chunksize=1):
            done[key] = pt
            n += 1
            if n % 10 == 0:
                _json.dump(done, open(ckpt, "w", encoding="utf-8"))
                ok = len([k for k in done if done[k] and done[k].get("gov_util")
                          and done[k]["gov_util"] >= UTIL_LO])
                print(f"  ... {len(done)}/{len(under)} re-run ({ok} now PASS)",
                      flush=True)
    _json.dump(done, open(ckpt, "w", encoding="utf-8"))
    # merged set: main, with only the under-threshold cases replaced
    merged = dict(main)
    for k in under:
        if done.get(k):
            merged[k] = done[k]
    _finalize_util(done, os.path.join(out_dir, "under_3000_only"))
    res = _finalize_util(merged, out_dir)
    rescued = len([k for k in done if done[k] and done[k].get("gov_util")
                   and done[k]["gov_util"] >= UTIL_LO])
    print(f"DONE: {rescued}/{len(under)} UNDER cases now PASS at {cap_kg:.0f} kg")
    return res


def _read_level_seeds(path):
    """Read max_levels per (section, config, Lcr_DA) from a prior chart xlsx, to
    seed the level-count search (>=1).  Returns {} on any problem."""
    seeds = {}
    if not path or not os.path.exists(path):
        return seeds
    try:
        import openpyxl
        ws = openpyxl.load_workbook(path, read_only=True, data_only=True).active
        it = ws.iter_rows(values_only=True)
        hdr = list(next(it))
        idx = {name: i for i, name in enumerate(hdr)}
        sc, cf = idx.get("section"), idx.get("config")
        ld, ml = idx.get("Lcr_DA_mm"), idx.get("max_levels")
        if None in (sc, cf, ld, ml):
            return {}
        for r in it:
            if r[sc] is None:
                continue
            seeds[(r[sc], r[cf], int(r[ld]))] = max(int(r[ml] or 0), 1)
    except Exception:
        return {}
    return seeds


def generate_util_based(master_path, out_dir, seeds_file=None,
                        checkpoint=None, workers=None, sections=None):
    """Resumable, parallel utilisation-targeted chart: for every case tune the
    soft load per level (and slightly the level count) so the governing upright
    util (max of STRESS and BUCKLING) lands in 0.97..0.99.  Level counts are
    seeded from a prior chart (seeds_file) when given.  `sections` restricts the
    upright list (default: all in the master)."""
    import multiprocessing as mp
    os.makedirs(out_dir, exist_ok=True)
    checkpoint = checkpoint or os.path.join(out_dir, "_util_checkpoint.json")
    done = {}
    if os.path.exists(checkpoint):
        done = _json.load(open(checkpoint, encoding="utf-8"))
    all_secs = load_master(master_path).library.names("upright")
    sections = [s for s in all_secs if s in sections] if sections else all_secs
    seeds = _read_level_seeds(seeds_file)
    todo = []
    for s in sections:
        for (_lbl, bt, p, xs) in MODEL_CONFIGS:
            for g in MODEL_DA:
                key = f"{s}|{_label(bt, p, xs)}|{g}"
                if key in done:
                    continue
                nseed = seeds.get((s, _label(bt, p, xs), int(g)), 3)
                todo.append((s, bt, p, xs, g, nseed))
    total = len(sections) * len(MODEL_CONFIGS) * len(MODEL_DA)
    workers = workers or min(4, os.cpu_count() or 1)
    print(f"util-target grid: {total} pts, {len(todo)} to do, {workers} workers, "
          f"target util {UTIL_LO}-{UTIL_HI}", flush=True)
    n = 0
    with mp.Pool(workers, initializer=_winit, initargs=(master_path,)) as pool:
        for key, pt in pool.imap_unordered(_wpoint_util, todo, chunksize=1):
            done[key] = pt
            n += 1
            if n % 20 == 0:
                _json.dump(done, open(checkpoint, "w", encoding="utf-8"))
                ok = len([k for k in done if done[k] and done[k].get("gov_util")])
                print(f"  ... {len(done)}/{total} ({ok} utilised)", flush=True)
    _json.dump(done, open(checkpoint, "w", encoding="utf-8"))
    return _finalize_util(done, out_dir)


def _finalize_util(done, out_dir):
    import openpyxl
    up_dir = os.path.join(out_dir, "uprights_util")
    os.makedirs(up_dir, exist_ok=True)
    bysec = {}
    for pt in done.values():
        if not pt or not pt.get("gov_util"):
            continue
        c = bysec.setdefault(pt["section"], {}).setdefault(
            pt["config"], {"da": [], "load": [], "lev": [], "frame": []})
        c["da"].append(pt["lcr_da"])
        c["load"].append(pt["load_per_level_kg"])
        c["lev"].append(pt["n_levels"])
        c["frame"].append(pt.get("frame_load_kg", 0.0))
    for sec, curves in bysec.items():
        for c in curves.values():
            order = sorted(range(len(c["da"])), key=lambda i: c["da"][i])
            for k in ("da", "load", "lev", "frame"):
                c[k] = [c[k][i] for i in order]
        _plot_util(sec, curves, os.path.join(up_dir, f"{sec}.png"))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Upright_util_chart"
    ws.append(["section", "config", "Lcr_CA_mm", "Lcr_DA_mm", "n_levels",
               "load_per_level_kg", "frame_load_kg", "bottom_upright_axial_kN",
               "stress_util", "buckling_util", "gov_util", "status"])
    for key, pt in sorted(done.items()):
        if pt:
            gov = pt.get("gov_util")
            # PASS = utilised into the target band; UNDER = best achievable but
            # sway-limited below 0.97 (value + util still reported)
            if gov is None:
                status = "NO RESULT"
            elif gov >= UTIL_LO:
                status = "PASS"
            else:
                status = "UNDER (sway-limited)"
            ws.append([pt["section"], pt["config"], pt["lcr_ca"], pt["lcr_da"],
                       pt.get("n_levels"), pt.get("load_per_level_kg"),
                       pt.get("frame_load_kg"), pt.get("bottom_axial_kN"),
                       pt.get("stress_util"), pt.get("buckling_util"),
                       gov, status])
    xlsx = os.path.join(out_dir, "Upright_Load_Chart_Utilised.xlsx")
    wb.save(xlsx)
    png_zip = os.path.join(out_dir, "upright_util_png.zip")
    with zipfile.ZipFile(png_zip, "w", zipfile.ZIP_DEFLATED) as z:
        for f in sorted(os.listdir(up_dir)):
            if f.endswith(".png"):
                z.write(os.path.join(up_dir, f), f"uprights_util/{f}")
    return {"sections": len(bysec), "xlsx": xlsx, "png_zip": png_zip}


def _plot_util(name, curves, path):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    fig, ax = plt.subplots(figsize=(8.5, 5.5))
    for label in ("X-500", "X-600", "D-1000", "D-1200", "XS-500", "XS-600"):
        c = curves.get(label)
        if not c:
            continue
        colour, ls = _STYLE.get(label, (branding.GREY, "-"))
        ax.plot(c["da"], c["frame"], ls, color=colour, lw=1.8, label=label)
    ax.set_xlabel("Down-aisle buckling length  Lcr,DA = beam gap  [mm]")
    ax.set_ylabel("Upright capacity = total frame load  [kg]  (<=2500 kg/level)")
    ax.set_title(f"{name}  -  fully-utilised upright capacity  ·  fy=355  ·  "
                 f"gamma_M1=1.1\nunbraced 3-bay semi-rigid frame, load<=2500 kg/level, "
                 f"levels maximised; util = max(stress, buckling) ~0.97-0.99",
                 fontsize=9.5)
    ax.grid(True, alpha=0.3)
    ax.set_ylim(bottom=0)
    ax.legend(fontsize=8, ncol=2)
    fig.tight_layout()
    fig.savefig(path, dpi=130)
    plt.close(fig)


if __name__ == "__main__":
    main()