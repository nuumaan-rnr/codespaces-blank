"""Importer for the user's section master workbook (.xlsx).

Expected sheets (units as in the workbook: cm, kN):
  UPRIGHT_MASTER : Section | Description | dims | Aeff cm2 | Iyy cm4 |
                   Izz cm4 | iyy | izz | Weff,y cm3 | Weff,z cm3 |
                   fy kN/cm2 | H depth | t wall cm | ...
  BEAM_MASTER    : # | SECTION | h mm | b mm | t mm | I cm4 | Wel cm3 |
                   fy kN/cm2 | M_Rd kNcm | EI
  BRACING_MASTER : transposed (property rows x section columns) with
                   Area cm2, Iyy/Izz cm4, Zyy/Zzz cm3, IT cm4, Fy kN/cm2,
                   thickness mm
  BASE_STIFFNESS : per upright, table of N (kN) vs k_b (kNcm/rad) vs
                   M_Rd (kNcm)  - the EN 15512 load-dependent floor
                   connection stiffness

Everything is converted to the app's N/mm unit set.

Axis mapping (see rack15512.model): the workbook's Iyy is the MAJOR axis
(engaged by down-aisle bending of uprights / gravity bending of beams) and
maps to the model's local-z properties (Iz, Welz); the workbook's Izz maps
to Iy/Wely.

Derived values (not in the workbook):
  * upright torsion constant J ~ A*t^2/3 (open thin-walled estimate),
  * beam minor-axis I/W, area and J from the RHS h x b x t dimensions.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

from .library import SectionLibrary
from .model import CrossSection

# unit factors -> N, mm
CM2 = 1.0e2     # cm^2  -> mm^2
CM3 = 1.0e3     # cm^3  -> mm^3
CM4 = 1.0e4     # cm^4  -> mm^4
KNCM2 = 10.0    # kN/cm^2 -> MPa
KNCM = 1.0e4    # kN*cm -> N*mm
KN = 1.0e3      # kN -> N


@dataclass
class MasterWorkbook:
    """Parsed master: a SectionLibrary plus the base-stiffness tables."""

    library: SectionLibrary
    # upright name -> sorted [(N [N], k_b [N*mm/rad], M_Rd [N*mm]), ...]
    base_tables: Dict[str, List[Tuple[float, float, float]]] = field(
        default_factory=dict)
    # section name -> fy [MPa] from the workbook
    fy: Dict[str, float] = field(default_factory=dict)

    def base_stiffness(self, upright: str, N: float) -> Tuple[float, float]:
        """Floor-connection stiffness and moment resistance for axial load
        N [N], linearly interpolated in the upright's table (clamped at the
        table ends)."""
        table = self.base_tables.get(upright)
        if not table:
            raise KeyError(f"No BASE_STIFFNESS table for upright '{upright}'")
        if N <= table[0][0]:
            return table[0][1], table[0][2]
        if N >= table[-1][0]:
            return table[-1][1], table[-1][2]
        for (n0, k0, m0), (n1, k1, m1) in zip(table, table[1:]):
            if n0 <= N <= n1:
                t = (N - n0) / (n1 - n0)
                return k0 + t * (k1 - k0), m0 + t * (m1 - m0)
        return table[-1][1], table[-1][2]    # unreachable


def load_master(path: str, role_hint: Optional[str] = None) -> MasterWorkbook:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("reading .xlsx masters requires openpyxl "
                          "(pip install openpyxl)") from None
    wb = openpyxl.load_workbook(path, data_only=True)
    if _is_rfem_persheet(wb):
        import os as _os
        hint = role_hint or _infer_role(_os.path.basename(path))
        return _load_rfem_persheet(wb, hint)
    sections: Dict[str, CrossSection] = {}
    fy_map: Dict[str, float] = {}

    if "UPRIGHT_MASTER" in wb.sheetnames:
        for s, fy in _parse_uprights(wb["UPRIGHT_MASTER"]):
            sections[s.name] = s
            fy_map[s.name] = fy
    if "BEAM_MASTER" in wb.sheetnames:
        for s, fy in _parse_beams(wb["BEAM_MASTER"]):
            sections[s.name] = s
            fy_map[s.name] = fy
    if "BRACING_MASTER" in wb.sheetnames:
        for s, fy in _parse_bracings(wb["BRACING_MASTER"]):
            sections[s.name] = s
            fy_map[s.name] = fy
    if not sections:
        raise ValueError(f"Master '{path}': no recognised sheets "
                         "(UPRIGHT_MASTER / BEAM_MASTER / BRACING_MASTER)")

    base = {}
    if "BASE_STIFFNESS" in wb.sheetnames:
        base = _parse_base_stiffness(wb["BASE_STIFFNESS"])

    return MasterWorkbook(library=SectionLibrary(sections),
                          base_tables=base, fy=fy_map)


# --------------------------------------------------------------------- RFEM
# Per-section property export (one upright per sheet) with columns
# Description | Symbol | Value | Unit | Comment, e.g. RFEM/RSTAB "effective
# section" sheets.  Carries the FULL property spectrum (shear areas, warping
# constant, shear centre, section moduli, buckling curves).

def _persheet_layout(ws):
    """Locate a Description/Symbol/Value section sheet's header, tolerating a
    blank leading row and/or a blank leading column; returns a dict of the
    header row and the Description/Symbol/Value/Comment column indices, or None
    when the sheet is not in that format."""
    for r in range(1, min(ws.max_row, 8) + 1):
        labels = {}
        for c in range(1, min(ws.max_column, 10) + 1):
            v = str(ws.cell(r, c).value or "").strip().lower()
            if v in ("description", "symbol", "value", "unit", "comment"):
                labels[v] = c
        if "symbol" in labels and "value" in labels:
            return {"row": r,
                    "desc": labels.get("description", labels["symbol"] - 1),
                    "sym": labels["symbol"], "val": labels["value"],
                    "comment": labels.get("comment", labels["value"] + 2)}
    return None


def _is_rfem_persheet(wb) -> bool:
    """True for an RFEM per-section export: no standard *_MASTER sheets and at
    least one sheet uses the Description/Symbol/Value layout."""
    if any(s in wb.sheetnames for s in
           ("UPRIGHT_MASTER", "BEAM_MASTER", "BRACING_MASTER")):
        return False
    return any(_persheet_layout(wb[s]) is not None for s in wb.sheetnames)


# role inference: from a master/file name keyword, else from a section name.
# 'others' is the catch-all role for products that are not uprights, beams or
# bracings (rails, connectors, shuttle parts, ...).
def _infer_role(text) -> Optional[str]:
    t = str(text or "").upper()
    if "UPRIGHT" in t:
        return "upright"
    if "BRACING" in t or "BRACE" in t:
        return "bracing"
    if "BEAM" in t:
        return "beam"
    if "OTHER" in t:
        return "others"
    return None


def _role_from_section(name) -> str:
    u = str(name or "").upper()
    if u.startswith("UP"):
        return "upright"
    if u.startswith("RHS"):
        return "beam"
    if u.startswith("1C") or u.startswith("C"):
        return "bracing"
    if ("DRIVEIN" in u or "SHUTTLE" in u or "RAIL" in u or u.startswith("CONN")):
        return "others"
    return "others"


def _persheet_props(ws) -> Dict[str, Tuple]:
    """{symbol: (value, comment)} for the first occurrence of each symbol."""
    lay = _persheet_layout(ws)
    if lay is None:
        return {}
    d: Dict[str, Tuple] = {}
    for r in range(lay["row"] + 1, ws.max_row + 1):
        sym = ws.cell(r, lay["sym"]).value
        if sym is None:
            continue
        sym = str(sym).strip()
        if sym and sym not in d:
            d[sym] = (ws.cell(r, lay["val"]).value,
                      ws.cell(r, lay["comment"]).value)
    return d


# perforated-upright calibration: the gross thin-wall thickness 2A/U (A = area,
# U = full double-sided perimeter) underestimates the actual sheet gauge by
# ~18% because the slots/perforations inflate U; this factor recovers the gauge
# (validated against the known 1.6/2.0/2.5/3.0/3.5 mm gauges of the sample set).
_UPRIGHT_PERF_FACTOR = 1.18


def _wall_thickness(A_mm2: float, U_cm: Optional[float]) -> Optional[float]:
    """Estimate the upright sheet gauge [mm] from the section area A [mm^2] and
    the RFEM perimeter U [cm]: t ~ 1.18 * 2A / U (perforation-corrected).  Used
    only to resolve the beam connector stiffness by upright thickness; it is an
    estimate and remains editable on the section."""
    if not (A_mm2 and U_cm and U_cm > 0):
        return None
    return round(_UPRIGHT_PERF_FACTOR * A_mm2 / (5.0 * U_cm), 2)


def _dist_mm(comment) -> Optional[float]:
    """Extract the fibre distance from an 'in distance 45.0 mm' comment [mm]."""
    import re
    m = re.search(r"(-?\d+(?:\.\d+)?)\s*mm", str(comment or ""))
    return float(m.group(1)) if m else None


def _load_rfem_persheet(wb, role_hint: Optional[str] = None) -> MasterWorkbook:
    """Build a MasterWorkbook from a per-sheet RFEM property export (one section
    per sheet).  Works for uprights, beams, bracings and other (rail/connector)
    sections - the role is `role_hint` (from the file/master name) when given,
    else inferred per section from its name.

    RFEM axes are swapped to the model convention (see module docstring): RFEM
    Iy (major) -> local Iz; RFEM Iz (minor) -> local Iy.  The shear areas,
    section moduli and buckling curves follow the same swap.  fy is recovered
    from the plastic axial force Npl,d / A (rounded to 5 MPa)."""
    sections: Dict[str, CrossSection] = {}
    fy_map: Dict[str, float] = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        p = _persheet_props(ws)
        if "A" not in p or "Iy" not in p:
            continue
        name = sheet.strip()
        if name.lower().endswith("-eff"):
            name = name[:-4].strip()
        role = role_hint or _role_from_section(name)

        def val(sym):
            try:
                return float(p.get(sym, (None,))[0])
            except (TypeError, ValueError):
                return None

        def dist(sym):
            return _dist_mm(p.get(sym, (None, None))[1])

        def modulus(maxsym, minsym):
            vals = [abs(x) for x in (val(maxsym), val(minsym)) if x is not None]
            return min(vals) * CM3 if vals else None

        def curve(sym):
            x = str(p.get(sym, ("b",))[0] or "b").strip().lower()
            return x if x in ("a0", "a", "b", "c", "d") else "b"

        A = (val("A") or 0.0) * CM2
        Iz = (val("Iy") or 0.0) * CM4               # RFEM major -> local z
        Iy = (val("Iz") or 0.0) * CM4               # RFEM minor -> local y
        J = (val("J") or 0.0) * CM4
        Avy = val("Az") * CM2 if val("Az") else None   # swap: local-y shear
        Avz = val("Ay") * CM2 if val("Ay") else None
        Iw = val("I@v,M") * 1.0e6 if val("I@v,M") else None   # cm6 -> mm6
        y0 = val("yM") * 10.0 if val("yM") is not None else None
        Welz = modulus("Sy,max", "Sy,min")          # about RFEM y -> local z
        Wely = modulus("Sz,max", "Sz,min")          # about RFEM z -> local y
        dh = (dist("Sy,max") or 0.0) - (dist("Sy,min") or 0.0) or None
        bw = (dist("Sz,max") or 0.0) - (dist("Sz,min") or 0.0) or None
        npl = val("Npl,d")                           # kN
        fy = round((npl * KN / A) / 5.0) * 5.0 if (npl and A > 0) else 350.0
        # wall thickness (gauge) for uprights, estimated from area and perimeter
        # so the beam connector stiffness can resolve by the upright thickness
        t = _wall_thickness(A, val("U")) if role == "upright" else None

        sections[name] = CrossSection(
            name=name, material="steel", role=role,
            A=A, Iy=Iy, Iz=Iz, J=J or (A * 4.0),    # tiny J fallback if missing
            Wely=Wely or 1.0, Welz=Welz or 1.0, t=t,
            A_eff=A, Wy_eff=Wely, Wz_eff=Welz,
            buckling_curve_y=curve("BCz/v"), buckling_curve_z=curve("BCy/u"),
            Avy=Avy, Avz=Avz, It_gross=J or None, Iw_gross=Iw, y0=y0,
            Iy_gross=Iy, Iz_gross=Iz, depth_h=dh, width_b=bw,
            description=f"RFEM full properties ({sheet})")
        fy_map[name] = fy
    if not sections:
        raise ValueError("no upright property sheets found in the RFEM export")
    return MasterWorkbook(library=SectionLibrary(sections),
                          base_tables={}, fy=fy_map)


def load_upright_properties(path: str) -> MasterWorkbook:
    """Public loader for an RFEM per-sheet upright property export."""
    import openpyxl
    return _load_rfem_persheet(openpyxl.load_workbook(path, data_only=True),
                               role_hint="upright")


def _norm_section(name) -> str:
    """Normalise a section name for matching across sheets (drop spaces /
    case): 'RHS 60x40x1.2' == 'RHS60X40X1.2'."""
    import re
    return re.sub(r"\s+", "", str(name or "")).upper()


def parse_section_geometry(path: str) -> Dict[str, Dict]:
    """Parse a geometry table (UPRIGHT_MASTER / BEAM_MASTER with columns like
    Section, Thickness (mm), H depth, Width, edge 1/2, or Sec Height/Depth/Thick)
    into {normalised name: {'t', 'depth_h', 'width_b', 'e1', 'e2'}} in mm.

    Only the explicit sheet-gauge thickness and edge distances are taken as
    authoritative; depth/width are filled best-effort (the RFEM full-property
    import already sets them)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    out: Dict[str, Dict] = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        cols: Dict = {}
        hdr = None
        for r in range(1, min(ws.max_row, 10) + 1):
            low = {c: str(ws.cell(r, c).value or "").strip().lower()
                   for c in range(1, ws.max_column + 1)}
            if any(v == "section" for v in low.values()):
                hdr = r
                for c, v in low.items():
                    if v == "section":
                        cols["section"] = c
                    elif "thick" in v:           # 'Thickness (mm)'/'Sec Thick'
                        cols["t"] = c
                    elif "h depth" in v:         # cm -> mm
                        cols["depth_h"] = (c, 10.0)
                    elif v == "sec height":
                        cols["depth_h"] = (c, 1.0)
                    elif v.startswith("width") or v == "sec depth":
                        cols["width_b"] = (c, 1.0)   # values are mm despite label
                    elif "edge 1" in v or v == "e1":
                        cols["e1"] = c
                    elif "edge 2" in v or v == "e2":
                        cols["e2"] = c
                break
        if hdr is None or "section" not in cols or "t" not in cols:
            continue
        for r in range(hdr + 1, ws.max_row + 1):
            nm = ws.cell(r, cols["section"]).value
            t = _num(ws.cell(r, cols["t"]).value)
            if not nm or not t:
                continue
            entry: Dict = {"t": t}                # thickness already in mm
            for key in ("depth_h", "width_b"):
                if key in cols:
                    c, fac = cols[key]
                    v = _num(ws.cell(r, c).value)
                    if v:
                        entry[key] = v * fac
            for key in ("e1", "e2"):
                if key in cols:
                    v = _num(ws.cell(r, cols[key]).value)
                    if v:
                        entry[key] = v
            out[_norm_section(nm)] = entry
    return out


def parse_beam_stiffness(path: str) -> Dict[str, Dict]:
    """Parse a beam-to-upright connector-stiffness sheet into
    {normalised section name: {'kb': [[upl_mm, k N*mm/rad], ...], 'm_rd': N*mm}}.

    The sheet has a 'Section' column, an optional 'M_Rd' column [kNcm] and one
    or more 'Kb @ UPL <t>' columns [kNcm/rad] (t = upright wall thickness)."""
    import openpyxl
    import re
    wb = openpyxl.load_workbook(path, data_only=True)
    out: Dict[str, Dict] = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        cols: Dict = {}
        header_row = None
        for r in range(1, min(ws.max_row, 12) + 1):
            vals = {c: str(ws.cell(r, c).value or "").strip()
                    for c in range(1, ws.max_column + 1)}
            low = {c: v.lower() for c, v in vals.items()}
            if (any(v == "section" for v in low.values())
                    and any("kb" in v for v in low.values())):
                header_row = r
                for c, v in low.items():
                    if v == "section":
                        cols["section"] = c
                    elif v.startswith("m_rd") or v.startswith("mrd"):
                        cols["m_rd"] = c
                    elif "kb" in v:
                        mm = re.search(r"upl\s*([\d.]+)", v)
                        if mm:
                            cols.setdefault("kb", []).append(
                                (float(mm.group(1)), c))
                break
        if header_row is None or "section" not in cols or not cols.get("kb"):
            continue
        for r in range(header_row + 1, ws.max_row + 1):
            nm = ws.cell(r, cols["section"]).value
            if not nm:
                continue
            kb = []
            for upl, c in sorted(cols["kb"]):
                v = _num(ws.cell(r, c).value)
                if v:
                    kb.append([upl, v * KNCM])          # kNcm/rad -> N*mm/rad
            if not kb:
                continue
            entry: Dict = {"kb": kb}
            if "m_rd" in cols:
                mrd = _num(ws.cell(r, cols["m_rd"]).value)
                if mrd:
                    entry["m_rd"] = mrd * KNCM          # kNcm -> N*mm
            out[_norm_section(nm)] = entry
    return out


def _rows(ws) -> List[list]:
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _num(v, default: Optional[float] = None) -> Optional[float]:
    if v is None or str(v).strip() == "":
        return default
    return float(v)


def _geom_upright(name, desc, header, cells):
    """Build an upright CrossSection from a GEOMETRY-only row (no A/I/W columns):
    overall depth, width and wall thickness -> gross channel properties.  Used
    when the UPRIGHT_MASTER sheet carries only dimensions; the exact perforated
    properties can be merged later (MasterStore.merge_stiffness)."""
    from .cf_sections import lipped_channel

    def col(*keys, fac=1.0):
        for idx, h in enumerate(header):
            if any(k in h for k in keys) and idx < len(cells):
                v = _num(cells[idx])
                return v * fac if v is not None else None
        return None

    t = col("thickness (mm", "thickness(mm") or col("t wall", fac=10.0) or 1.6
    depth = col("h depth", fac=10.0) or col("length") or 90.0   # mm
    width = col("width") or 61.0                                 # mm
    e1 = col("edge 1", "e1")
    e2 = col("edge 2", "e2")
    cs = lipped_channel(name, depth, width, 0.0, t)   # plain channel (no lip)
    # our convention: local z = strong/down-aisle (the larger inertia)
    if cs.Iy >= cs.Iz:
        Iz, Iy, Welz, Wely = cs.Iy, cs.Iz, cs.Wely, cs.Welz
    else:
        Iz, Iy, Welz, Wely = cs.Iz, cs.Iy, cs.Welz, cs.Wely
    return CrossSection(
        name=name, material="steel", role="upright",
        A=cs.A, Iy=Iy, Iz=Iz, J=cs.J, Wely=Wely, Welz=Welz,
        A_eff=cs.A, Wy_eff=Wely, Wz_eff=Welz,
        buckling_curve_y="b", buckling_curve_z="b",
        t=t, e1=e1, e2=e2, depth_h=depth, width_b=width,
        It_gross=cs.J, Iy_gross=Iy, Iz_gross=Iz,
        description=f"{desc} (geometry-derived gross properties)")


def _parse_uprights(ws):
    out = []
    header = None
    geom_only = False
    # optional gross torsion/warping/shear-centre columns for FT buckling,
    # found by header text (any position): IT/J, Iw, y0, Iyy_g, Izz_g
    cols = {}
    for r in _rows(ws):
        cells = r[1:]                       # data starts in column B
        if header is None:
            if cells and str(cells[0]).strip() == "Section":
                header = [str(c or "").strip().lower() for c in cells]
                # geometry-only sheet: no area / second-moment columns present
                geom_only = not any(
                    ("aeff" in h or "iyy" in h or "izz" in h
                     or h.strip() in ("a", "a (cm2)", "a cm2")) for h in header)
                for idx, h in enumerate(header):
                    if ("it" in h or h.startswith("j")) and "cm4" in h:
                        cols["It"] = idx
                    elif "iw" in h or "warping" in h:
                        cols["Iw"] = idx
                    elif h.startswith("y0") or "shear cen" in h:
                        cols["y0"] = idx
            continue
        if not cells or not cells[0]:
            continue

        if geom_only:
            name = str(cells[0]).strip()
            desc = str(cells[1] or "").strip() if len(cells) > 1 else ""
            out.append((_geom_upright(name, desc, header, cells), 355.0))
            continue

        def opt(key, fac):
            idx = cols.get(key)
            if idx is None or idx >= len(cells):
                return None
            v = _num(cells[idx])
            return v * fac if v else None

        name = str(cells[0]).strip()
        desc = str(cells[1] or "").strip()
        A_eff = _num(cells[5]) * CM2
        Iz = _num(cells[6]) * CM4           # workbook Iyy = major = local z
        Iy = _num(cells[7]) * CM4           # workbook Izz = minor = local y
        Wz_eff = _num(cells[10]) * CM3      # Weff,y -> Welz
        Wy_eff = _num(cells[11]) * CM3      # Weff,z -> Wely
        fy = _num(cells[12], 35.0) * KNCM2
        t = _num(cells[14], 0.2) * 10.0     # t wall cm -> mm
        It = opt("It", CM4)
        J = It if It else A_eff * t * t / 3.0     # open-section estimate
        out.append((CrossSection(
            name=name, material="steel", role="upright",
            A=A_eff, Iy=Iy, Iz=Iz, J=J,
            Wely=Wy_eff, Welz=Wz_eff,
            A_eff=A_eff, Wy_eff=Wy_eff, Wz_eff=Wz_eff,
            buckling_curve_y="b", buckling_curve_z="b",
            t=t, e1=_num(cells[15]), e2=_num(cells[16]),
            depth_h=_num(cells[2]), width_b=_num(cells[3]),
            It_gross=It, Iw_gross=opt("Iw", 1.0e6),   # cm6 -> mm6
            y0=opt("y0", 10.0),                       # cm -> mm
            Iy_gross=Iy, Iz_gross=Iz,
            description=f"{desc} (A=Aeff)"), fy))
    return out


def _parse_beams(ws):
    out = []
    header = None
    k_col = mrd_col = loos_col = None
    k_fac = mrd_fac = KNCM            # default workbook units: kNcm(/rad)
    loos_fac = 1.0e-3                 # default: mrad
    for r in _rows(ws):
        cells = r[1:]
        if header is None:
            if cells and str(cells[0]).strip() == "#":
                header = [str(c or "").strip().lower() for c in cells]
                # optional per-beam connector columns, found by header text:
                #   'Connector k (kNcm/rad)' or '(kNm/rad)'
                #   'Connector M_Rd (kNcm)' or '(kNm)'
                #   'Connector looseness (mrad)' or '(rad)' / 'phi_l'
                for idx, h in enumerate(header):
                    if "connector" in h and ("stiff" in h or " k" in h
                                             or h.startswith("k")):
                        k_col = idx
                        k_fac = 1.0e6 if "knm" in h else KNCM
                    elif "connector" in h and ("m_rd" in h or "mrd" in h
                                               or "m rd" in h):
                        mrd_col = idx
                        mrd_fac = 1.0e6 if "knm" in h else KNCM
                    elif "looseness" in h or "phi_l" in h:
                        loos_col = idx
                        loos_fac = 1.0 if "(rad" in h else 1.0e-3
            continue
        if not cells or not cells[1]:
            continue
        name = str(cells[1]).strip()
        h = _num(cells[2])                  # mm
        b = _num(cells[3])
        t = _num(cells[4])
        fy = _num(cells[7], 27.0) * KNCM2 if len(cells) > 7 else 355.0

        def opt(col, fac):
            if col is None or col >= len(cells):
                return None
            v = _num(cells[col])
            return v * fac if v else None

        # minor axis, area and J from the RHS h x b x t geometry
        hi, bi = h - 2 * t, b - 2 * t
        A = h * b - hi * bi
        Iy = (h * b**3 - hi * bi**3) / 12.0
        Wely = 2.0 * Iy / b
        # major axis I/W: from the sheet when given, else from the geometry
        Iz_sheet = _num(cells[5]) if len(cells) > 5 else None
        Iz = Iz_sheet * CM4 if Iz_sheet else (b * h**3 - bi * hi**3) / 12.0
        Welz_sheet = _num(cells[6]) if len(cells) > 6 else None
        Welz = Welz_sheet * CM3 if Welz_sheet else 2.0 * Iz / h
        hm, bm = h - t, b - t               # closed thin-wall torsion
        J = 2.0 * t * (hm * bm) ** 2 / (hm + bm)
        out.append((CrossSection(
            name=name, material="steel", role="beam",
            A=A, Iy=Iy, Iz=Iz, J=J, Wely=Wely, Welz=Welz,
            t=t, depth_h=h, width_b=b,
            buckling_curve_y="b", buckling_curve_z="b",
            connector_k=opt(k_col, k_fac),
            connector_m_rd=opt(mrd_col, mrd_fac),
            connector_looseness=opt(loos_col, loos_fac),
            description=f"RHS {h:.0f}x{b:.0f}x{t:.1f} "
                        "(minor axis/J from geometry)"), fy))
    return out


_BRACE_ROW_KEYS = (
    ("section area", "A", CM2),
    ("iyy", "Iz", CM4),                     # workbook Iyy = major -> local z
    ("izz", "Iy", CM4),
    ("zyy", "Welz", CM3),
    ("zzz", "Wely", CM3),
    ("it", "J", CM4),
    ("fy", "fy", KNCM2),
    ("fu", "fu", KNCM2),
    ("thk", "t", 1.0),                      # mm
    ("end dist. e1", "e1", 1.0),            # mm
    ("end dist. e2", "e2", 1.0),
    ("iw", "Iw", 1.0e6),                    # cm6 -> mm6
    ("y0", "y0", 10.0),                     # cm -> mm
)


def _parse_bracings(ws):
    rows = _rows(ws)
    if not rows:
        return []
    names_raw = [str(v).strip() for v in rows[0][1:] if v is not None]
    names: List[str] = []
    for n in names_raw:                     # de-duplicate repeated names
        names.append(n if n not in names else f"{n} #{names.count(n) + 1}")
    props: Dict[str, List[float]] = {}
    for r in rows[1:]:
        label = str(r[0] or "").strip().lower()
        for key, attr, factor in _BRACE_ROW_KEYS:
            if label.startswith(key):
                props[attr] = [(_num(v, 0.0) or 0.0) * factor
                               for v in r[1:1 + len(names)]]
    out = []
    for k, name in enumerate(names):
        def p(attr, default=0.0):
            vals = props.get(attr)
            return vals[k] if vals and k < len(vals) else default
        out.append((CrossSection(
            name=name, material="steel", role="bracing",
            A=p("A"), Iy=p("Iy"), Iz=p("Iz"), J=max(p("J"), 1.0),
            Wely=max(p("Wely"), 1.0), Welz=max(p("Welz"), 1.0),
            buckling_curve_y="c", buckling_curve_z="c",
            t=p("t", None), e1=p("e1", None), e2=p("e2", None),
            fu=p("fu", None) or None,
            It_gross=p("J", None) or None, Iw_gross=p("Iw", None) or None,
            y0=p("y0", None), Iy_gross=p("Iy") or None,
            Iz_gross=p("Iz") or None,
            description="cold-formed C brace"), p("fy", 270.0)))
    return out


def _parse_base_stiffness(ws) -> Dict[str, List[Tuple[float, float, float]]]:
    tables: Dict[str, List[Tuple[float, float, float]]] = {}
    current: Optional[str] = None
    for r in _rows(ws):
        cells = r[1:]
        if not cells:
            continue
        first = str(cells[0] or "").strip()
        if first and not _is_number(first):
            if first.lower().startswith("n (kn"):
                continue                     # header row of a block
            if first.upper().startswith("UP"):
                current = first
                tables[current] = []
            continue
        if current and _is_number(first):
            N = float(first) * KN
            k = _num(cells[1])
            m = _num(cells[2])
            if k is not None:
                tables[current].append((N, k * KNCM, (m or 0.0) * KNCM))
    return {k: sorted(v) for k, v in tables.items() if v}


def _is_number(s: str) -> bool:
    try:
        float(s)
        return True
    except ValueError:
        return False
