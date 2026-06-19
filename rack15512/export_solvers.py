"""Export a RackModel to neutral solver input files so a project can be re-run
in RSTAB/RFEM or STAAD.Pro instead of (or to cross-check) the OpenSeesPy engine.

ADDITIVE only - reads a RackModel, writes input decks; no engine changes.

  * to_staad(model, path)      -> STAAD.Pro .std text input
  * to_rstab_xlsx(model, path) -> RSTAB/RFEM table .xlsx (Nodes/Members/Sections/
                                  Materials/Supports/Hinges/Loads), the same
                                  tabular layout RSTAB imports from Excel.

Axis convention: this app uses Z up (gravity = -Z).  STAAD.Pro SPACE uses Y up,
so coordinates are mapped (X, Y, Z)_staad = (X, Z, Y)_app and gravity is -Y.
RSTAB keeps the app's axes (Z up); loads are written with gravity in -Z.

Both decks carry the geometry, sections (as prismatic A/I/J), materials, the
nodal supports (with rotational spring constants), the beam-end connector hinges
(rotational springs), and the characteristic pallet + dead load case.  Load
combinations, imperfections and the 2nd-order setting are emitted as comments /
a note so the engineer applies the same EN 15512 combinations in the target
solver (they differ per code and are documented in the report).
"""

from __future__ import annotations

import math
from typing import Dict, List

from .model import RackModel


def _members_of(model: RackModel):
    return sorted(model.members.values(), key=lambda m: m.id)


def _node_map(model: RackModel) -> Dict[int, int]:
    """1-based node id remap (STAAD/RSTAB joints must be >= 1; app uses 0-based)."""
    return {nid: i + 1 for i, nid in enumerate(sorted(model.nodes))}


def _udl_load_case(model: RackModel):
    """Collect the characteristic gravity member UDLs [N/mm] (down = -Z) from the
    'pallets' + 'dead' load cases if present, keyed by member id."""
    q: Dict[int, float] = {}
    for lc in model.load_cases.values():
        if lc.name not in ("pallets", "dead"):
            continue       # skip pattern/placement/accidental variants
        for ml in getattr(lc, "member_loads", []):
            q[ml.member] = q.get(ml.member, 0.0) + getattr(ml, "qz", 0.0)
    return q


# ----------------------------------------------------------------- STAAD .std
def to_staad(model: RackModel, path: str) -> str:
    """Write a STAAD.Pro .std input deck (mm, N).  Returns the path."""
    L: List[str] = []
    L.append("STAAD SPACE")
    L.append("START JOB INFORMATION")
    L.append(f"ENGINEER DATE {model.name}")
    L.append("END JOB INFORMATION")
    L.append("* Exported from Racks & Rollers (EN 15512 app). App axes: Z up.")
    L.append("* STAAD SPACE uses Y up -> coords mapped (x,y,z)=(X,Z,Y); gravity=-Y.")
    L.append("UNIT MMS NEWTON")
    nmap = _node_map(model)

    # joints (Y up): staad_y = app_z, staad_z = app_y
    L.append("JOINT COORDINATES")
    for n in sorted(model.nodes.values(), key=lambda n: n.id):
        L.append(f"{nmap[n.id]} {n.x:.4f} {n.z:.4f} {n.y:.4f};")

    L.append("MEMBER INCIDENCES")
    for m in _members_of(model):
        L.append(f"{m.id} {nmap[m.node_i]} {nmap[m.node_j]};")

    # materials (one Steel per distinct E/G)
    L.append("DEFINE MATERIAL START")
    done = set()
    for mat in model.materials.values():
        nm = mat.name.replace(" ", "_")
        if nm in done:
            continue
        done.add(nm)
        L.append(f"ISOTROPIC {nm}")
        L.append(f"E {mat.E}")
        L.append(f"POISSON {getattr(mat, 'nu', 0.3)}")
        L.append("DENSITY 7.85e-08")
        L.append(f"G {mat.G}")
    L.append("END DEFINE MATERIAL")

    # prismatic section per member (AX/IX/IY/IZ in mm)
    L.append("MEMBER PROPERTY AMERICAN")
    for m in _members_of(model):
        s = model.section_of(m)
        a = s.A * m.area_factor
        L.append(f"{m.id} PRIS AX {a:.3f} IX {s.J:.1f} IY {s.Iy:.1f} IZ {s.Iz:.1f}")

    # material assignment per member
    L.append("CONSTANTS")
    for m in _members_of(model):
        nm = model.material_of(m).name.replace(" ", "_")
        L.append(f"MATERIAL {nm} MEMB {m.id}")

    # supports (translations fixed; rotational springs from support.ry etc.)
    L.append("SUPPORTS")
    for sup in model.supports:
        parts = [f"{sup.node}"]
        # STAAD FIXED BUT releases free DOFs; spring via KFX.. KMX..; map app
        # (ux,uy,uz,rx,ry,rz) -> staad (FX,FZ?,FY,..). Keep translations fixed.
        # Use a generic: translations fixed, rotational springs where given.
        kmx = sup.ry if isinstance(sup.ry, (int, float)) else 0.0   # app ry = down-aisle
        # STAAD: FIXED BUT MX MY MZ <springs>; we fix translations, spring rotations
        spr = []
        # app ry (about global Y / down-aisle bending) -> STAAD MY (about staad Y=app Z)?
        # Map app rotational DOFs to STAAD: app rx->MX, app ry->MZ(staad), app rz->MY
        def kk(v):
            return v if isinstance(v, (int, float)) and v > 0 else None
        rx, ry, rz = kk(sup.rx), kk(sup.ry), kk(sup.rz)
        rel = []
        if rx is None: rel.append("MX")
        else: spr.append(f"KMX {rx:.0f}")
        # app ry (down-aisle, about app Y) -> STAAD MZ (about staad Z = app Y)
        if ry is None: rel.append("MZ")
        else: spr.append(f"KMZ {ry:.0f}")
        if rz is None: rel.append("MY")
        else: spr.append(f"KMY {rz:.0f}")
        line = f"{nmap[sup.node]} FIXED BUT " + " ".join(rel + spr)
        L.append(line.strip())

    # beam-end connector hinges: partial moment release with rotational spring
    # (app hinge.rz about member local z = beam strong axis -> STAAD MZ)
    rel_lines = []
    for m in _members_of(model):
        for end, h in (("START", m.hinge_i), ("END", m.hinge_j)):
            if h is None:
                continue
            kz = getattr(h, "rz", None)
            if kz and kz > 0:
                rel_lines.append(f"{m.id} {end} KMZ {kz:.0f}")
            elif kz == 0:
                rel_lines.append(f"{m.id} {end} MZ")
    if rel_lines:
        L.append("MEMBER RELEASE")
        L.extend(rel_lines)

    # characteristic gravity load case (pallets+dead) as member UDL in -Y (staad)
    q = _udl_load_case(model)
    L.append("LOAD 1 LOADTYPE LIVE  TITLE PALLET + DEAD (characteristic)")
    L.append("MEMBER LOAD")
    for mid, qz in q.items():
        if abs(qz) > 1e-9:
            # app qz [N/mm] downward (-Z) -> STAAD GY (global Y up) negative
            L.append(f"{mid} UNI GY {qz:.5f}")
    L.append("LOAD 2 LOADTYPE DEAD  TITLE SELFWEIGHT")
    L.append("SELFWEIGHT Y -1.0")
    L.append("* EN 15512 ULS combos (apply in STAAD): 1.3 DL + 1.4 LL (+ placement),")
    L.append("* plus a sway imperfection and a 2nd-order (P-Delta) analysis:")
    L.append("LOAD COMB 101 1.3DL + 1.4LL")
    L.append("1 1.4 2 1.3")
    L.append("PERFORM ANALYSIS PRINT STATICS CHECK")
    L.append("* For 2nd order use:  PERFORM ANALYSIS  with  DEFINE PDELTA / PDELTA")
    L.append("FINISH")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(L) + "\n")
    return path


# --------------------------------------------------------------- RSTAB .xlsx
def to_rstab_xlsx(model: RackModel, path: str) -> str:
    """Write an RSTAB/RFEM table workbook (Nodes/Materials/Cross-Sections/
    Members/Nodal Supports/Member Hinges/Load Case) for Excel import."""
    import openpyxl
    wb = openpyxl.Workbook()

    nmap = _node_map(model)
    ws = wb.active
    ws.title = "1.1 Nodes"
    ws.append(["No.", "X [mm]", "Y [mm]", "Z [mm]"])
    for n in sorted(model.nodes.values(), key=lambda n: n.id):
        ws.append([nmap[n.id], n.x, n.y, n.z])

    ws = wb.create_sheet("1.2 Materials")
    ws.append(["No.", "Description", "E [N/mm2]", "G [N/mm2]", "nu", "fy [N/mm2]"])
    mats = {m.name: m for m in model.materials.values()}
    midx = {nm: i + 1 for i, nm in enumerate(mats)}
    for nm, m in mats.items():
        ws.append([midx[nm], nm, m.E, m.G, getattr(m, "nu", 0.3), m.fy])

    ws = wb.create_sheet("1.3 Cross-Sections")
    ws.append(["No.", "Description", "Material No.", "A [mm2]", "Iy [mm4]",
               "Iz [mm4]", "J [mm4]", "Wely [mm3]", "Welz [mm3]"])
    secs = {s.name: s for s in model.sections.values()}
    sidx = {nm: i + 1 for i, nm in enumerate(secs)}
    for nm, s in secs.items():
        ws.append([sidx[nm], nm, midx.get(s.material, 1), s.A, s.Iy, s.Iz,
                   s.J, s.Wely, s.Welz])

    ws = wb.create_sheet("1.7 Members")
    ws.append(["No.", "Start Node", "End Node", "Cross-Section No.",
               "Start Hinge", "End Hinge", "Member Set"])
    # collect distinct hinges
    hinges: List = []
    def hinge_id(h):
        if h is None:
            return ""
        for i, hh in enumerate(hinges):
            if (hh.rx, hh.ry, hh.rz) == (h.rx, h.ry, h.rz):
                return i + 1
        hinges.append(h)
        return len(hinges)
    for m in _members_of(model):
        ws.append([m.id, nmap[m.node_i], nmap[m.node_j],
                   sidx.get(model.section_of(m).name, 1),
                   hinge_id(m.hinge_i), hinge_id(m.hinge_j), m.member_set])

    ws = wb.create_sheet("1.4 Member Hinges")
    ws.append(["No.", "phi-x [Nmm/rad]", "phi-y [Nmm/rad]", "phi-z [Nmm/rad]",
               "M_Rd,z [Nmm]", "comment"])
    for i, h in enumerate(hinges):
        ws.append([i + 1, h.rx or "free", h.ry or "free", h.rz or "free",
                   getattr(h, "m_rd_z", None), "beam-end connector"])

    ws = wb.create_sheet("1.8 Nodal Supports")
    ws.append(["On Nodes", "uX", "uY", "uZ", "phi-X [Nmm/rad]",
               "phi-Y [Nmm/rad]", "phi-Z [Nmm/rad]"])
    def cell(v):
        if v is True:
            return "fixed"
        if not v:
            return "free"
        return v
    for sup in model.supports:
        ws.append([nmap[sup.node], cell(sup.ux), cell(sup.uy), cell(sup.uz),
                   cell(sup.rx), cell(sup.ry), cell(sup.rz)])

    ws = wb.create_sheet("LC1 Member Loads")
    ws.append(["On Members", "Type", "Direction", "p [N/mm]", "comment"])
    q = _udl_load_case(model)
    for mid, qz in q.items():
        if abs(qz) > 1e-9:
            ws.append([mid, "Force", "Z (down)", qz, "pallet+dead characteristic"])

    ws = wb.create_sheet("README")
    for row in [
        ["Exported from Racks & Rollers (EN 15512 app) for RSTAB/RFEM import."],
        ["Axes: Z up, gravity -Z (RSTAB default global Z can be set down)."],
        ["Cross-sections are prismatic A/Iy/Iz/J; assign the real SHAPE-THIN /"],
        ["RRO sections in RSTAB if available for exact warping/shear."],
        ["Apply EN 15512 ULS combos (1.3DL+1.4LL+..), sway imperfection and a"],
        ["2nd-order (P-Delta) analysis in RSTAB to reproduce the app's checks."],
    ]:
        ws.append(row)

    wb.save(path)
    return path


def export_project(model_json: str, out_dir: str, name: str = "model") -> dict:
    """Load a saved project model.json and write both solver decks."""
    import os
    from . import io_json
    os.makedirs(out_dir, exist_ok=True)
    model = io_json.load(model_json)
    std = to_staad(model, os.path.join(out_dir, f"{name}.std"))
    xlsx = to_rstab_xlsx(model, os.path.join(out_dir, f"{name}_RSTAB.xlsx"))
    return {"staad": std, "rstab_xlsx": xlsx, "nodes": len(model.nodes),
            "members": len(model.members)}
