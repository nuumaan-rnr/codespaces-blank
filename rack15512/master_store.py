"""In-app section-master database.

Instead of re-reading an Excel workbook on every run, masters are stored
*inside the system* as JSON and can be created, updated and deleted.  A
master holds the section library (uprights / beams / bracings with their
full properties), the per-upright floor-connection (BASE_STIFFNESS) tables,
and the per-section yield strengths.

    masters/
      <master-id>/
        master.json        # sections + base tables + fy + metadata

Workflow:
    * import an existing Excel master ONCE  -> stored as JSON
    * thereafter edit / add / delete sections in place (UI, CLI or API)
    * builds reference a stored master by id; .to_workbook() hands the
      builder the same MasterWorkbook it used to get from the .xlsx
"""

from __future__ import annotations

import datetime as _dt
import json
import os
import re
from dataclasses import asdict, dataclass, field, fields
from typing import Any, Dict, List, Optional, Tuple

from .library import SectionLibrary
from .master_xlsx import MasterWorkbook, load_master
from .model import CrossSection


def _now() -> str:
    return _dt.datetime.now().isoformat(timespec="seconds")


def _slug(text: str) -> str:
    return re.sub(r"[^a-zA-Z0-9]+", "-", text.strip().lower()).strip("-") \
        or "master"


def _section_to_dict(s: CrossSection) -> Dict[str, Any]:
    return asdict(s)


def _section_from_dict(d: Dict[str, Any]) -> CrossSection:
    known = {f.name for f in fields(CrossSection)}
    return CrossSection(**{k: v for k, v in d.items() if k in known})


def builtin_others_master() -> "StoredMaster":
    """The built-in 'others' master holding the custom drive-in sections:
    'Drivein Rail' (the depth rail) and 'drive in connector' (the cantilever
    arm with its connector stiffness, RSTAB Konsole jZ = 1.0e6 N*mm/rad)."""
    from .drive_in import _rstab_arm, _rstab_rail
    rail = _rstab_rail()
    rail.name, rail.role = "Drivein Rail", "beam"
    rail.description = "Drive-in depth rail (RSTAB DRIVE-IN RAIL 2.5)"
    conn = _rstab_arm()
    conn.name, conn.role = "drive in connector", "beam"
    conn.description = "Drive-in cantilever arm + connector (RSTAB Konsole)"
    conn.connector_k = 1.0e6                   # jZ = 100 kN.cm/rad
    conn.connector_looseness = 0.0
    m = StoredMaster(id="others", name="others",
                     description="Custom sections (drive-in rail & connector)")
    m.upsert_section(rail, fy=350.0)
    m.upsert_section(conn, fy=350.0)
    return m


@dataclass
class StoredMaster:
    """A self-contained, editable section master."""

    id: str
    name: str
    company: str = ""                 # owning company (sections are company-specific)
    description: str = ""
    created: str = field(default_factory=_now)
    updated: str = field(default_factory=_now)
    # section name -> serialised CrossSection
    sections: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    # upright name -> [[N, k_b, M_Rd], ...] (N units throughout: N, N*mm/rad)
    base_tables: Dict[str, List[List[float]]] = field(default_factory=dict)
    # section name -> fy [MPa]
    fy: Dict[str, float] = field(default_factory=dict)

    # ---- conversions ----------------------------------------------------
    def to_workbook(self) -> MasterWorkbook:
        secs = {name: _section_from_dict(d)
                for name, d in self.sections.items()}
        base = {k: [tuple(row) for row in v]
                for k, v in self.base_tables.items()}
        return MasterWorkbook(library=SectionLibrary(secs),
                              base_tables=base, fy=dict(self.fy))

    @property
    def library(self) -> SectionLibrary:
        return self.to_workbook().library

    def roles(self) -> List[str]:
        return sorted({d.get("role", "") for d in self.sections.values()
                       if d.get("role")})

    def names(self, role: Optional[str] = None) -> List[str]:
        return sorted(n for n, d in self.sections.items()
                      if role is None or d.get("role") == role)

    # ---- section CRUD ---------------------------------------------------
    def upsert_section(self, sec: CrossSection,
                       fy: Optional[float] = None) -> None:
        self.sections[sec.name] = _section_to_dict(sec)
        if fy is not None:
            self.fy[sec.name] = fy
        self.updated = _now()

    def update_fields(self, name: str, **changes) -> None:
        if name not in self.sections:
            raise KeyError(f"section '{name}' not in master '{self.id}'")
        valid = {f.name for f in fields(CrossSection)}
        for k, v in changes.items():
            if k == "fy":
                self.fy[name] = float(v)
            elif k in valid:
                self.sections[name][k] = v
            else:
                raise KeyError(f"unknown section field '{k}'")
        self.updated = _now()

    def delete_section(self, name: str) -> None:
        self.sections.pop(name, None)
        self.fy.pop(name, None)
        self.base_tables.pop(name, None)
        self.updated = _now()

    def set_base_table(self, upright: str,
                       table: List[Tuple[float, float, float]]) -> None:
        self.base_tables[upright] = [list(row) for row in table]
        self.updated = _now()

    # ---- (de)serialisation ----------------------------------------------
    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)

    @classmethod
    def from_dict(cls, d: Dict[str, Any]) -> "StoredMaster":
        known = {f.name for f in fields(cls)}
        return cls(**{k: v for k, v in d.items() if k in known})

    @classmethod
    def from_workbook(cls, mw: MasterWorkbook, mid: str, name: str,
                      description: str = "") -> "StoredMaster":
        sections = {n: _section_to_dict(s)
                    for n, s in mw.library.sections.items()}
        base = {k: [list(row) for row in v]
                for k, v in mw.base_tables.items()}
        return cls(id=mid, name=name, description=description,
                   sections=sections, base_tables=base, fy=dict(mw.fy))


class MasterStore:
    """Filesystem-backed store of section masters."""

    def __init__(self, root: str = "masters"):
        self.root = root
        os.makedirs(self.root, exist_ok=True)
        self._companies_file = os.path.join(self.root, "_companies.json")

    # ---- companies (the company master) ---------------------------------
    def companies(self) -> List[str]:
        """Registered company names, unioned with any referenced by masters."""
        names = set()
        if os.path.isfile(self._companies_file):
            try:
                with open(self._companies_file, encoding="utf-8") as f:
                    names.update(json.load(f))
            except Exception:
                pass
        names.update(m.company for m in self.list() if m.company)
        return sorted(n for n in names if n)

    def add_company(self, name: str) -> None:
        name = (name or "").strip()
        if not name:
            return
        names = set(self.companies())
        names.add(name)
        with open(self._companies_file, "w", encoding="utf-8") as f:
            json.dump(sorted(names), f, indent=2)

    def delete_company(self, name: str) -> None:
        names = [n for n in self.companies() if n != name]
        with open(self._companies_file, "w", encoding="utf-8") as f:
            json.dump(names, f, indent=2)

    def by_company(self) -> Dict[str, List["StoredMaster"]]:
        """Stored masters grouped by company (key '' = unassigned)."""
        out: Dict[str, List["StoredMaster"]] = {}
        for m in self.list():
            out.setdefault(m.company or "", []).append(m)
        return out

    def _file(self, mid: str) -> str:
        return os.path.join(self.root, mid, "master.json")

    def exists(self, mid: str) -> bool:
        return os.path.isfile(self._file(mid))

    def list(self) -> List[StoredMaster]:
        out = []
        for name in sorted(os.listdir(self.root)):
            if os.path.isfile(self._file(name)):
                out.append(self.load(name))
        return out

    def load(self, mid: str) -> StoredMaster:
        with open(self._file(mid), encoding="utf-8") as f:
            return StoredMaster.from_dict(json.load(f))

    def save(self, master: StoredMaster) -> None:
        os.makedirs(os.path.join(self.root, master.id), exist_ok=True)
        with open(self._file(master.id), "w", encoding="utf-8") as f:
            json.dump(master.to_dict(), f, indent=2)

    def delete(self, mid: str) -> None:
        import shutil
        d = os.path.join(self.root, mid)
        if os.path.isdir(d):
            shutil.rmtree(d)

    def merge_stiffness(self, mid: str, path: str) -> Tuple[int, int]:
        """Merge supplementary data into an existing master IN PLACE, matched by
        section name: a beam-connector sheet sets connector_k_by_upl / M_Rd (and
        a default connector_k); a geometry sheet sets the section thickness t and
        edge distances (and depth/width when missing); a BASE_STIFFNESS sheet
        adds per-upright base tables.  Returns (sections_updated, base_added)."""
        from .master_xlsx import (_norm_section, _parse_base_stiffness,
                                  parse_beam_stiffness, parse_section_geometry)
        m = self.load(mid)
        by_norm: Dict[str, str] = {}
        for nm in m.sections:
            by_norm.setdefault(_norm_section(nm), nm)

        touched = set()
        for key, entry in parse_beam_stiffness(path).items():
            target = by_norm.get(key)
            if target is None:
                continue
            sec = m.sections[target]
            if entry.get("kb"):
                sec["connector_k_by_upl"] = entry["kb"]
                mid_row = sorted(entry["kb"])[len(entry["kb"]) // 2]
                sec["connector_k"] = mid_row[1]        # default (middle UPL)
            if entry.get("m_rd"):
                sec["connector_m_rd"] = entry["m_rd"]
            touched.add(target)

        # geometry / thickness (the explicit sheet gauge replaces any estimate)
        for key, entry in parse_section_geometry(path).items():
            target = by_norm.get(key)
            if target is None:
                continue
            sec = m.sections[target]
            sec["t"] = entry["t"]
            for k in ("e1", "e2"):
                if entry.get(k) is not None:
                    sec[k] = entry[k]
            for k in ("depth_h", "width_b"):           # fill only if missing
                if entry.get(k) and not sec.get(k):
                    sec[k] = entry[k]
            touched.add(target)
        updated = len(touched)

        base_added = 0
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            if "BASE_STIFFNESS" in wb.sheetnames:
                for up, rows in _parse_base_stiffness(wb["BASE_STIFFNESS"]).items():
                    tgt = by_norm.get(_norm_section(up), up)
                    m.base_tables[tgt] = [list(r) for r in rows]
                    base_added += 1
        except Exception:
            pass

        m.updated = _now()
        self.save(m)
        return updated, base_added

    def _unique_id(self, base: str) -> str:
        mid, k = base, 2
        while self.exists(mid):
            mid = f"{base}-{k}"
            k += 1
        return mid

    # ---- creation -------------------------------------------------------
    def create(self, name: str, description: str = "") -> StoredMaster:
        m = StoredMaster(id=self._unique_id(_slug(name)), name=name,
                         description=description)
        self.save(m)
        return m

    def ensure_builtin(self) -> None:
        """Create the built-in 'others' master (Drivein Rail + drive in
        connector) if it is not already in the store."""
        if not self.exists("others"):
            self.save(builtin_others_master())

    def import_xlsx(self, path: str, name: Optional[str] = None,
                    description: str = "", company: str = "") -> StoredMaster:
        """Import an Excel (or CSV/JSON) master ONCE into the store.  A
        company name is mandatory (sections are company-specific)."""
        company = (company or "").strip()
        if not company:
            raise ValueError("a company name is required to import a master")
        if path.lower().endswith((".xlsx", ".xlsm")):
            from .master_xlsx import _infer_role
            # role hint for RFEM per-sheet exports comes from the import name
            # (e.g. BEAM_MASTER / BRACING_MASTER / UPRIGHT_PROPERTIES)
            mw = load_master(path, role_hint=_infer_role(name or path))
        else:
            lib = SectionLibrary.from_file(path)
            fy = {s.name: 250.0 for s in lib.sections.values()}
            mw = MasterWorkbook(library=lib, base_tables={}, fy=fy)
        name = name or os.path.splitext(os.path.basename(path))[0]
        m = StoredMaster.from_workbook(mw, self._unique_id(_slug(name)),
                                       name, description)
        m.company = company
        self.add_company(company)
        self.save(m)
        return m
