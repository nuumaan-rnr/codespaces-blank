"""Post-process the model-based upright load chart into a clean, monotonic
(downward) "linear" load chart.

The raw model output (rack15512.loadcharts util chart) is noisy: the integer
level count and the near-critical limit produce occasional UP-SPIKES in the
capacity-vs-buckling-length curve, and a few points slightly overshoot the
target utilisation.  Physically the upright capacity must DECREASE monotonically
with the unsupported (down-aisle buckling) length.  This module:

  1. normalises every point to a utilisation of UTIL_TARGET (0.97): the reported
     capacity is axial * UTIL_TARGET / gov_util, so all lengths are on one util;
  2. removes up-spikes by linear interpolation between the bracketing points
     (50 mm grid) and enforces a non-increasing curve per (section, config);
  3. enforces the configuration ordering D <= X <= XS at every length;
  4. enforces the perforation ordering: an odd-numbered upright (extra holes)
     cannot carry more than the even-numbered one below it (UP0003<=UP0002, ...)
     up to UP0021.

Outputs a cleaned workbook + per-section before/after charts + a validation
report.  Pure data processing - no FEA.
"""

from __future__ import annotations

import os
import zipfile
from typing import Dict, List, Tuple

import openpyxl

from . import branding

UTIL_TARGET = 0.97
# perforation pairs: odd (with extra holes) must be <= even (no extra holes)
ODD_EVEN_PAIRS = [(f"UP{e:04d}", f"UP{e+1:04d}") for e in range(2, 21, 2)]  # ..UP0021
_STYLE = {"X-600": (branding.TEAL, "-"), "D-1200": (branding.GREY, "--"),
          "XS-600": (branding.TEAL_LIGHT, "-.")}


def _clean_series(L: List[int], cap: List[float]) -> List[float]:
    """Return a non-increasing capacity series: up-spikes are replaced by linear
    interpolation between the bracketing trend points; a residual increase is
    clamped.  L is ascending (uniform 50 mm), cap the capacity at each L."""
    c = list(cap)
    n = len(c)
    i = 1
    while i < n:
        if c[i] > c[i - 1] + 1e-9:                 # up-spike begins at i
            j = i
            while j < n and c[j] > c[i - 1] + 1e-9:  # next point back on trend
                j += 1
            if j < n:                              # interpolate i..j-1
                for k in range(i, j):
                    t = (L[k] - L[i - 1]) / (L[j] - L[i - 1])
                    c[k] = c[i - 1] + t * (c[j] - c[i - 1])
                i = j
            else:                                  # spike to the end: extrapolate
                # use the average downward slope of the clean head
                slope = 0.0
                if i - 1 >= 1 and L[i - 1] != L[0]:
                    slope = (c[i - 1] - c[0]) / (L[i - 1] - L[0])  # <= 0
                for k in range(i, n):
                    c[k] = max(c[i - 1] + slope * (L[k] - L[i - 1]), 0.0)
                break
        else:
            i += 1
    for i in range(1, n):                          # safety: strictly non-increasing
        if c[i] > c[i - 1]:
            c[i] = c[i - 1]
    return c


def load_chart(in_xlsx: str) -> Dict[Tuple[str, str], Dict[int, dict]]:
    """Read the util-chart workbook -> {(section,config): {Lcr_DA: row dict}}.
    row dict carries the raw axial, util, and the 0.97-normalised capacity."""
    ws = openpyxl.load_workbook(in_xlsx, read_only=True, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    h = list(rows[0])
    idx = {name: i for i, name in enumerate(h)}
    out: Dict[Tuple[str, str], Dict[int, dict]] = {}
    for r in rows[1:]:
        sec, cfg = r[idx["section"]], r[idx["config"]]
        gov = r[idx["gov_util"]]
        frame = r[idx["frame_load_kg"]]            # total load carried (kg) =
        #                          n_levels * load/level - comparable across D/X/XS
        if sec is None or gov in (None, 0) or not frame:
            continue
        L = int(r[idx["Lcr_DA_mm"]])
        cap097 = frame * UTIL_TARGET / gov         # total capacity at util = 0.97
        out.setdefault((sec, cfg), {})[L] = {
            "lcr_ca": r[idx["Lcr_CA_mm"]], "frame_raw": frame, "gov_util": gov,
            "n_levels": r[idx["n_levels"]], "load_kg": r[idx["load_per_level_kg"]],
            "axial_raw": r[idx["bottom_upright_axial_kN"]], "cap_raw": cap097}
    return out


def clean(in_xlsx: str, out_dir: str) -> dict:
    data = load_chart(in_xlsx)
    # 1) per-curve spike removal + monotonic
    clean_cap: Dict[Tuple[str, str], Dict[int, float]] = {}
    for key, byL in data.items():
        Ls = sorted(byL)
        caps = _clean_series(Ls, [byL[L]["cap_raw"] for L in Ls])
        clean_cap[key] = {L: caps[i] for i, L in enumerate(Ls)}

    # 2) ordering D <= X <= XS at each length (clip the lower ones down)
    sections = sorted({s for (s, _c) in clean_cap})
    Lgrid = sorted({L for byL in data.values() for L in byL})
    for sec in sections:
        D = clean_cap.get((sec, "D-1200")); X = clean_cap.get((sec, "X-600"))
        XS = clean_cap.get((sec, "XS-600"))
        for L in Lgrid:
            xs = XS.get(L) if XS else None
            x = X.get(L) if X else None
            d = D.get(L) if D else None
            # anchor on the well-sampled X curve: D clipped down to X, XS raised
            # up to X (the stiffener can only add capacity, never reduce it)
            if x is not None:
                if d is not None and d > x:
                    D[L] = x                         # D <= X
                if xs is not None and xs < x:
                    XS[L] = x                        # XS >= X
            elif xs is not None and d is not None and d > xs:
                D[L] = xs                            # no X curve: D <= XS

    # 3) perforation ordering: odd (holes) <= even, per config/length
    for even, odd in ODD_EVEN_PAIRS:
        for cfg in ("X-600", "D-1200", "XS-600"):
            ce = clean_cap.get((even, cfg)); co = clean_cap.get((odd, cfg))
            if not ce or not co:
                continue
            for L in list(co):
                if L in ce and co[L] > ce[L]:
                    co[L] = ce[L]
    # re-enforce monotonic after clipping
    for key, byL in clean_cap.items():
        Ls = sorted(byL)
        caps = _clean_series(Ls, [byL[L] for L in Ls])
        clean_cap[key] = {L: caps[i] for i, L in enumerate(Ls)}

    os.makedirs(out_dir, exist_ok=True)
    png_dir = os.path.join(out_dir, "uprights_clean")
    os.makedirs(png_dir, exist_ok=True)
    for sec in sections:
        _plot(sec, data, clean_cap, os.path.join(png_dir, f"{sec}.png"))

    # workbook: cleaned capacity (+ raw for comparison)
    wb = openpyxl.Workbook(); wsr = wb.active; wsr.title = "Clean_load_chart"
    wsr.append(["section", "config", "Lcr_CA_mm", "Lcr_DA_mm",
                "frame_load_kg_util0.97", "raw_frame_load_kg", "raw_gov_util",
                "interpolated"])
    for (sec, cfg) in sorted(clean_cap):
        byL = data.get((sec, cfg), {})
        for L in sorted(clean_cap[(sec, cfg)]):
            cleaned = clean_cap[(sec, cfg)][L]
            raw = byL.get(L, {})
            interp = abs(cleaned - raw.get("cap_raw", cleaned)) > 0.05
            wsr.append([sec, cfg, raw.get("lcr_ca"), L, round(cleaned, 2),
                        round(raw.get("cap_raw", 0), 2),
                        round(raw.get("gov_util", 0), 3), "yes" if interp else ""])
    # validation sheet
    wsv = wb.create_sheet("Validation")
    wsv.append(["check", "result", "violations"])
    dxx = _check_ordering(clean_cap, sections, Lgrid)
    oe = _check_pairs(clean_cap)
    mono = _check_monotonic(clean_cap)
    wsv.append(["D <= X <= XS", "PASS" if not dxx else "FAIL",
                "; ".join(dxx[:5])])
    wsv.append(["odd(holes) <= even", "PASS" if not oe else "FAIL",
                "; ".join(oe[:5])])
    wsv.append(["monotonic non-increasing", "PASS" if not mono else "FAIL",
                "; ".join(mono[:5])])
    xlsx = os.path.join(out_dir, "Upright_Load_Chart_Linear.xlsx")
    wb.save(xlsx)
    zp = os.path.join(out_dir, "upright_clean_png.zip")
    with zipfile.ZipFile(zp, "w", zipfile.ZIP_DEFLATED) as z:
        for f in sorted(os.listdir(png_dir)):
            if f.endswith(".png"):
                z.write(os.path.join(png_dir, f), f"uprights_clean/{f}")
    return {"sections": len(sections), "xlsx": xlsx, "png_zip": zp,
            "D<=X<=XS": "PASS" if not dxx else f"FAIL({len(dxx)})",
            "odd<=even": "PASS" if not oe else f"FAIL({len(oe)})",
            "monotonic": "PASS" if not mono else f"FAIL({len(mono)})"}


def _check_ordering(cc, sections, Lgrid):
    bad = []
    for sec in sections:
        D = cc.get((sec, "D-1200")); X = cc.get((sec, "X-600"))
        XS = cc.get((sec, "XS-600"))
        for L in Lgrid:
            d = D.get(L) if D else None; x = X.get(L) if X else None
            xs = XS.get(L) if XS else None
            if d is not None and x is not None and d > x + 1e-6:
                bad.append(f"{sec} L{L}: D>{x:.1f}")
            if x is not None and xs is not None and x > xs + 1e-6:
                bad.append(f"{sec} L{L}: X>XS")
    return bad


def _check_pairs(cc):
    bad = []
    for even, odd in ODD_EVEN_PAIRS:
        for cfg in ("X-600", "D-1200", "XS-600"):
            ce = cc.get((even, cfg)); co = cc.get((odd, cfg))
            if not ce or not co:
                continue
            for L in co:
                if L in ce and co[L] > ce[L] + 1e-6:
                    bad.append(f"{odd}>{even} {cfg} L{L}")
    return bad


def _check_monotonic(cc):
    bad = []
    for key, byL in cc.items():
        Ls = sorted(byL)
        for i in range(1, len(Ls)):
            if byL[Ls[i]] > byL[Ls[i - 1]] + 1e-6:
                bad.append(f"{key} L{Ls[i]}")
    return bad


def _plot(sec, data, clean_cap, path):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    fig, ax = plt.subplots(figsize=(8.5, 5.5))
    for cfg in ("D-1200", "X-600", "XS-600"):
        cc = clean_cap.get((sec, cfg))
        if not cc:
            continue
        colour, ls = _STYLE.get(cfg, (branding.GREY, "-"))
        Ls = sorted(cc)
        raw = data.get((sec, cfg), {})
        ax.plot(Ls, [raw[L]["cap_raw"] for L in Ls], color=colour, lw=0.8,
                alpha=0.35)                          # raw (faint)
        ax.plot(Ls, [cc[L] for L in Ls], ls, color=colour, lw=2.0,
                label=f"{cfg} (clean)")
    ax.set_xlabel("Down-aisle buckling length  Lcr,DA  [mm]")
    ax.set_ylabel("Upright capacity = total frame load  [kg]  at util = 0.97")
    ax.set_title(f"{sec}  -  linear upright load chart (de-spiked, monotonic)\n"
                 f"faint = raw model;  solid = cleaned;  D <= X <= XS",
                 fontsize=9.5)
    ax.grid(True, alpha=0.3); ax.set_ylim(bottom=0)
    ax.legend(fontsize=8)
    fig.tight_layout(); fig.savefig(path, dpi=130); plt.close(fig)
