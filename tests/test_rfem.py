"""Tests for the RFEM importer and the validation against the SPR
reference model's exported results."""

import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pytest

from rack15512.combos import assemble, apply_ehf
from rack15512.engine.opensees import OpenSeesEngine
from rack15512.master_xlsx import load_master
from rack15512.model import DIRECTION_VECTORS, Combination
from rack15512.rfem_compare import (MemberRef, compare_results,
                                    read_rfem_results)
from rack15512.rfem_import import load_rfem

HERE = os.path.dirname(__file__)
DATA = os.path.join(HERE, "..", "examples", "SPR_CHECK_Data.xlsx")
MASTER = os.path.join(HERE, "..", "examples", "Master.xlsx")
needs_data = pytest.mark.skipif(not os.path.exists(DATA),
                                reason="SPR_CHECK_Data.xlsx not present")

# the importer splits member 711 for its concentrated load; its station
# extremes are no longer comparable 1:1 with the RFEM full-length member
SPLIT = (711, 529)


@pytest.fixture(scope="module")
def model():
    return load_rfem(DATA)


@needs_data
def test_rfem_import_structure(model):
    assert len(model.supports) == 16
    assert len(model.members) == 529          # 528 + 1 split
    assert model.imperfection.value() == pytest.approx(1.0 / 300.0)
    beams = [m for m in model.members.values() if m.hinge_i]
    assert len(beams) == 60
    assert beams[0].hinge_i.rz == pytest.approx(6573.0 * 1e4)
    trusses = [m for m in model.members.values() if m.mtype == "truss"]
    assert len(trusses) == 176
    # pinned bases in the export (nonlinear springs are not exported)
    assert model.supports[0].ry is False
    # Z flipped to Z-up
    assert min(n.z for n in model.nodes.values()) == 0.0
    assert max(n.z for n in model.nodes.values()) == pytest.approx(9050.0, abs=1.0)
    # self weight rebuilt from material density (RFEM total: 12.09 kN)
    lc1 = model.load_cases["LC1"]
    tot = sum(abs(ml.qz) * model.member_length(model.members[ml.member])
              for ml in lc1.member_loads)
    assert tot == pytest.approx(12.09e3, rel=0.01)
    # combinations with per-combo imperfection senses
    co1 = next(c for c in model.combinations if c.name.startswith("CO1 "))
    assert co1.imp_directions == ["+x"]
    co4 = next(c for c in model.combinations if c.name.startswith("CO4 "))
    assert co4.imp_directions == ["+y"]


@needs_data
def test_rfem_linear_lc2_matches(model):
    """RFEM solves load cases linearly; our first-order LC2 member forces
    must reproduce the export (validates geometry, hinges, supports,
    sections and loads in one shot)."""
    ref = {}
    import openpyxl
    wb = openpyxl.load_workbook(DATA, data_only=True, read_only=True)
    sheet = next(s for s in wb.sheetnames if s.startswith("LC2 - 4.1"))
    cur = None
    for r in wb[sheet].iter_rows(min_row=3, values_only=True):
        head = "" if r[0] is None else str(r[0]).strip()
        if head.isdigit():
            cur = ref.setdefault(int(head), MemberRef())
        if cur is None or r[3] is None:
            continue
        try:
            N, My, Mz = float(r[3]) * 1e3, float(r[7]) * 1e4, float(r[8]) * 1e4
        except (TypeError, ValueError):
            continue
        cur.N_min = min(cur.N_min, N)
        cur.N_max = max(cur.N_max, N)
        cur.Mz_absmax = max(cur.Mz_absmax, abs(My))
        cur.My_absmax = max(cur.My_absmax, abs(Mz))

    case = OpenSeesEngine().run_case(
        model, assemble(model, Combination("LC2", "SLS", {"LC2": 1.0})),
        name="LC2", combo="LC2", kind="SLS", order=1)
    assert case.converged
    comps = compare_results(model, [case], {"LC2": ref}, skip_members=SPLIT)
    diffs = sorted(c.rel_diff for c in comps)
    assert len(comps) > 300
    assert diffs[len(diffs) // 2] < 0.02          # median < 2%


@needs_data
def test_rfem_second_order_with_base_springs_matches():
    """The reference model's nonlinear combinations only reproduce when the
    load-dependent floor-connection springs from the master are restored;
    CO1 member forces then match RFEM's second-order results."""
    master = load_master(MASTER)
    model = load_rfem(DATA, master=master)
    assert isinstance(model.supports[0].ry, float)   # spring recovered
    rfem = read_rfem_results(DATA)
    combo = next(c for c in model.combinations if c.name.startswith("CO1 "))
    loads = apply_ehf(model, assemble(model, combo),
                      model.imperfection.value(), DIRECTION_VECTORS["+x"])
    case = OpenSeesEngine().run_case(model, loads, name="CO1", combo="CO1",
                                     kind="ULS", order=2, imp_direction="+x")
    assert case.converged
    comps = compare_results(model, [case], {"CO1": rfem["CO1"]},
                            skip_members=SPLIT)
    mo = sorted(c.rel_diff for c in comps if not c.quantity.startswith("N"))
    ax = sorted(c.rel_diff for c in comps if c.quantity.startswith("N"))
    assert ax[len(ax) // 2] < 0.01                # axials median < 1%
    assert mo[len(mo) // 2] < 0.03                # moments median < 3%


if __name__ == "__main__":
    sys.exit(pytest.main([__file__, "-v"]))
