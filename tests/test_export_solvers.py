"""Solver-export (STAAD .std + RSTAB .xlsx) smoke tests."""
import os, sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from rack15512 import export_solvers as ex
from rack15512.builder import RackConfig, build_rack, LevelSpec


def _model():
    return build_rack(RackConfig(module="single", n_bays=2,
        levels=[LevelSpec(gap=2000.0)], frame_height=2200.0))


def test_staad_deck(tmp_path):
    m = _model()
    p = ex.to_staad(m, str(tmp_path / "m.std"))
    txt = open(p).read()
    assert "STAAD SPACE" in txt and "JOINT COORDINATES" in txt
    assert "MEMBER INCIDENCES" in txt and "SUPPORTS" in txt
    assert "FINISH" in txt
    # 1-based joints: no "0 0.0000" joint line, lowest joint is 1
    assert "\n1 " in txt
    nmap = ex._node_map(m)
    assert min(nmap.values()) == 1


def test_rstab_xlsx(tmp_path):
    import openpyxl
    m = _model()
    p = ex.to_rstab_xlsx(m, str(tmp_path / "m.xlsx"))
    wb = openpyxl.load_workbook(p)
    assert "1.1 Nodes" in wb.sheetnames and "1.7 Members" in wb.sheetnames
    assert wb["1.7 Members"].max_row == len(m.members) + 1   # header + members


if __name__ == "__main__":
    import pytest
    sys.exit(pytest.main([__file__, "-v"]))
